import tkinter
from tkinter import ttk
from tkinter import filedialog
import re
import pandas as pd
from datetime import datetime, timedelta
import tkinter.messagebox
from PIL import Image, ImageTk
from openpyxl import load_workbook
from openpyxl.styles import Color
from tkcalendar import Calendar
from tkinterdnd2 import DND_FILES

import customizing
import sap_engine

def number_to_letters(num):
    result = ""
    while num > 0:
        num -= 1
        result = chr(65 + (num % 26)) + result
        num //= 26
    return result

def quit_window (window: tkinter.Tk):
    #quit_message = tkinter.messagebox.askyesnocancel(title='Warning', message='Are you sure you want to quit?')
    #if quit_message == tkinter.TRUE:
        window.destroy()

def home ():
    for widget in mainframe.frame.winfo_children():
        if type(widget) != tkinter.Menu:
            widget.destroy()
    MainRoot (root = mainroot.root, frame = mainframe.frame)

def start (program: str, frame: tkinter.Frame):
    for widget in frame.winfo_children():
        if type(widget) != tkinter.Menu:
            widget.destroy()
    if program == 'ebs_mt940':
        Ebs (frame = frame)
    elif program == 'iban':
        Iban (frame = frame)
    else:
        MigrationFile (frame = frame)

def tab_creation (frame, x: int, y: int, width: int = 1000, height: int = 800, anchor: str = 'nw', bordermode: str = 'inside'):
    tab = ttk.Notebook(frame)
    tab.place (x = x, y = y, width = width, height = height, anchor = anchor, bordermode = bordermode)
    return tab

def browse_file(entry_path: customizing.Entry):
    file_path = filedialog.askopenfilename(filetypes=[("Text files", "*.txt")])
    entry_path.entry.config (state = 'normal')
    entry_path.entry.delete(0, tkinter.END)  # Clear the entry widget
    entry_path.entry.insert(0, file_path)  # Insert the selected file path into the entry widget
    entry_path.entry.config (state = 'disabled')

def browse_file_xlsx(entry_path: customizing.Entry):
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    entry_path.entry.config(state = 'normal')
    entry_path.entry.delete(0, tkinter.END)  # Clear the entry widget
    entry_path.entry.insert(0, file_path)  # Insert the selected file path into the entry widget

def read_file(entry_path: customizing.Entry, text: customizing.TextEntry):
    file_path = entry_path.entry.get()
    text.drag_label.place_forget()
    try:
        with open(file_path, 'r') as file:
            content = file.read()
            text.text_entry.delete(1.0, tkinter.END)  # Clear the text widget
            text.text_entry.insert(tkinter.END, content)  # Insert the file content into the text widget
    except FileNotFoundError:
        text.text_entry.delete(1.0, tkinter.END)
        text.text_entry.insert(tkinter.END, "File not found.")

def read_file_xlsx(entry_path_str: str):
    file_path = entry_path_str
    with open(file_path, 'r') as file:
        content = pd.read_excel(file_path)

def create_excel_file(sheet_names_list: list, present_fields: list):
    # Ask the user to choose where to save the file
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

    if file_path:
        # Create a Pandas Excel writer using xlsxwriter as the engine
        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
            # Write each DataFrame to a different sheet
            for a in sheet_names_list:
                sheet_present_fields = []
                sheet_present_fields_tech = []
                for j in present_fields:
                    if j[0] == a[0]:
                        sheet_present_fields.append(j[1].replace("*", "").replace("+", ""))
                        sheet_present_fields_tech.append(j[4])

                # Create a DataFrame with the first row filled with field names and the second one with technical names
                df_sheet = pd.DataFrame([sheet_present_fields_tech], columns=sheet_present_fields)

                # Write DataFrame to Excel
                df_sheet.to_excel(writer, sheet_name=a[0], index=False)

                # Get the worksheet object
                worksheet = writer.sheets[a[0]]

                # Create a format for center alignment with a border
                center_border_format = writer.book.add_format({
                    'align': 'center',
                    'valign': 'vcenter',
                    'border': 1  # Adds a thin border around the cells
                })

                # Apply the format to the second row (technical names)
                for col_num, value in enumerate(sheet_present_fields):
                    worksheet.write(1, col_num, df_sheet.iloc[0, col_num], center_border_format)

                # Adjust column size based on the length of the column name
                for col_num, value in enumerate(sheet_present_fields, 0):
                    max_len = max(df_sheet[value].astype(str).apply(len).max(), len(value))
                    col_width = (max_len + 2) * 1.2  # Adjust the multiplier as needed
                    worksheet.set_column(col_num, col_num, col_width)

        print(f"Excel file saved to: {file_path}")

def migration_forget (sheet_list: list, sheet: str):
    for a in sheet_list:
        if sheet != 'No':
            a.sheet.checkbox.place_forget() 
        a.main_frame.frame.pack_forget()
        for b in a.field_list:
            b.label.label.place_forget()
            b.radiobutton_1.place_forget()
            b.radiobutton_2.place_forget()
            b.radiobutton_3.place_forget() 

def info (type: str):
    Info (type = type)

class Info:
    def __init__(
        self,
        type: str
        ):
        self.root = customizing.Root (root_title = 'HELP')
        self.root.root.state('normal')

        self.frame = customizing.Frame (
            root = self.root.root,
            pack_or_grid = 'P'
        )

        self.type_label = customizing.Label (
            frame = self.frame.frame,
            text = type,
            dimension = 24,
            x = 10,
            y = 10
        )

        self.main_label = customizing.Label (
            frame = self.frame.frame,
            text = '',
            x = 10,
            y = 70
        )

        if type == 'EBS MT940 info':
            self.main_label.label.config (text = customizing.ebs_mt940_text)
        elif type == 'IBAN info':
            self.main_label.label.config (text = customizing.iban_text)
        elif type == 'Migration File info':
            self.main_label.label.config (text = customizing.mf_text)

class MainRoot:
    def __init__(
        self,
        root: tkinter.Tk,
        frame: tkinter.Frame
        ):
        self.root = root

        self.frame = frame

        customizing.Label (
             frame = self.frame,
             text = 'SAP HELPER',
             dimension = 40,
             foreground = '#229F22',
             x = 650,
             y = 30,
        )
        customizing.Button (
            frame = self.frame,
            text = 'EBS MT940',
            command = lambda: start('ebs_mt940', self.frame),
            width = 15,
            height = 3,
            x = 10,
            y = 130
        )
        customizing.Button (
            frame = self.frame,
            text = 'IBAN',
            command = lambda: start('iban', self.frame),
            width = 15,
            height = 3,
            x = 10,
            y = 220
        )
        customizing.Button (
            frame = self.frame,
            text = 'Migration File',
            command = lambda: start('migration_file', self.frame),
            width = 15,
            height = 3,
            x = 10,
            y = 310
        )
        
class Ebs:
    def __init__(
        self,
        frame: tkinter.Frame
        ):
        def ebs_analysis ():
            self.file_path = self.entry_path.entry.get()

            ebs_eng = sap_engine.EbsEngine (content = self.text.text_entry.get("1.0", tkinter.END))

            for widget in frame.winfo_children():
                widget.destroy()

            self.headers = ['SWIFT', 'BANK ACCOUNT N°', 'START DATE', 'END DATE', 'CURRENCY', 'OPENING BALANCE', 'CLOSING BALANCE']
                
            self.rows = [[ebs_eng.swift, ebs_eng.bank_account_number, ebs_eng.start_date, ebs_eng.end_date, ebs_eng.currency, ebs_eng.opening_balance, ebs_eng.closing_balance],
                    ['', '', '', '', '', '', ''],
                    ['VALUE DATE', 'AMOUNT', 'BANK EXTERNAL TRANSACTION', 'BANK EXT TR DESCRIPTION']]
            
            
            for pos in ebs_eng.position_lst:
                self.rows.append(pos)
            self.rows.append ('')
            self.rows.append(['OPENING BALANCE', ebs_eng.opening_balance])
            self.rows.append(['TOTAL CREDIT', "{:.2f}".format(ebs_eng.total_credit)])
            self.rows.append(['TOTAL DEBIT', "{:.2f}".format(ebs_eng.total_debit)])
            self.rows.append(['CLOSING BALANCE', ebs_eng.closing_balance])
            self.rows.append(['CHECK', "{:.2f}".format(float(ebs_eng.opening_balance) + ebs_eng.total_credit + ebs_eng.total_debit - float(ebs_eng.closing_balance))])
        
            tree = customizing.Treeview (
                frame = self.frame,
                col_text = self.headers,
                width_list = [120, 120, 120, 120, 120, 120, 120],
                lst = self.rows
            )

            excel = customizing.Button (
                frame = self.frame,
                text = 'Export to Excel',
                command = lambda: tree.export_to_excel (file_path = self.file_path, file_name = "EBS MT940 export", ebs = "YES"),
                x = 300,
                y = 10
            )

            customizing.Label (
                frame = self.frame,
                text = f'EBS transactions: {len(ebs_eng.position_lst)}',
                weight = 'bold',
                x = 750,
                y = 15
            )

            check = float(ebs_eng.opening_balance) + ebs_eng.total_credit + ebs_eng.total_debit - float(ebs_eng.closing_balance)
            image_path = 'green_tick.png'
            if round(check, 2) != 0:
                image_path = 'red_cross.png'

            # Open the image using Pillow
            image = Image.open(image_path)
            # Convert the image to a format Tkinter supports
            icon_check = ImageTk.PhotoImage(image)

            label_with_icon = tkinter.Label(tree.frame, image=icon_check, text="CHECK: ", compound=tkinter.RIGHT, font = ('Calibri', 14, 'bold'), background = '#F0F8FF')
            label_with_icon.place (x = 600, y = 10)

            # Keep a reference to the image to prevent it from being garbage collected
            label_with_icon.image = icon_check


        self.frame = frame
        self.headers = []
        self.rows = []
        self.file_path = ''

        self.entry_path = customizing.Entry(
            frame = self.frame,
            width = 80,
            x = 150,
            y = 20
            )
        self.entry_path.entry.config(state = 'disabled')

        #button to upload the .txt file
        customizing.Button(
            frame = self.frame, 
            text = "Upload .txt file", 
            command = lambda: (browse_file (entry_path=self.entry_path), read_file (entry_path=self.entry_path, text=self.text)),
            x = 10,
            y = 10
            )

        #text field. It is automatically filled uploading the file. It's also possible to drag and drop the file
        self.text = customizing.TextEntry (
            frame = self.frame,
            text_height = 44,
            text_width = 100,
            height = 680,
            width = 800,
            entry_path = self.entry_path,
            x = 10,
            y = 70
            )

        #button to analyze the file content (a Trevieew will be opened)
        customizing.Button(
            frame = self.frame, 
            text = "Analysis", 
            command = ebs_analysis,
            x = 700,
            y = 10
            )

class Iban:
    def __init__(
        self,
        frame: tkinter.Frame
        ):
        def iban_analysis ():
            self.file_path = self.entry_path.entry.get()

            iban_eng = sap_engine.IbanEngine (content = self.text.text_entry.get("1.0", tkinter.END))

            for widget in frame.winfo_children():
                widget.destroy()
        
            self.headers = ['IBAN', 'BANK COUNTRY', 'BANK KEY', 'BANK ACCOUNT N°', 'BANK CONTROL KEY', 'SWIFT', 'NOTES']

            self.rows = []
            for pos in iban_eng.position_lst:
                self.rows.append(pos)

        
            tree = customizing.Treeview (
                frame = self.frame,
                col_text = self.headers,
                width_list = [200, 180, 200, 200, 200, 150, 800],
                lst = self.rows
            )

            excel = customizing.Button (
                frame = self.frame,
                text = 'Export to Excel',
                command = lambda: tree.export_to_excel (file_path = self.file_path, file_name = "IBAN"),
                x = 300,
                y = 10
            )

            customizing.Label (
                frame = self.frame,
                text = f'Total IBANs: {len(iban_eng.position_lst)}                        IBANs not analyzed: {iban_eng.iban_not_analyzed}',
                weight = 'bold',
                x = 500,
                y = 15
            )
            

        self.frame = frame
        self.headers = []
        self.rows = []
        self.file_path = ''

        self.entry_path = customizing.Entry(
            frame = self.frame,
            width = 80,
            x = 150,
            y = 20
            )
        self.entry_path.entry.config(state = 'disabled')

        #button to upload the .txt file
        customizing.Button(
            frame = self.frame, 
            text = "Upload .txt file", 
            command = lambda: (browse_file (entry_path=self.entry_path), read_file (entry_path=self.entry_path, text=self.text)),
            x = 10,
            y = 10
            )

        #text field. It is automatically filled uploading the file. It's also possible to drag and drop the file
        self.text = customizing.TextEntry (
            frame = self.frame,
            text_height = 44,
            text_width = 100,
            height = 680,
            width = 800,
            entry_path = self.entry_path,
            x = 10,
            y = 70
            )

        #button to analyze the file content (a Trevieew will be opened)
        customizing.Button(
            frame = self.frame, 
            text = "Analysis", 
            command = iban_analysis,
            x = 700,
            y = 10
            )

class MigrationFile:
    def __init__(
        self,
        frame: tkinter.Frame
        ):      
        self.frame = frame

        self.tab = ttk.Notebook (self.frame, height=680, width=1320)

        self.sheet_names = []

        self.sh_list = []

        self.sheet_names_list = []

        self.mode_key_fields = []

        self.present_fields = []

        self.error_list = []

        self.file_path = ''

        self.input_fields = []

        self.entry_path = customizing.Entry(
            frame = self.frame,
            width = 80,
            x = 150,
            y = 20
            )
        self.entry_path.entry.config(state = 'disabled')
        
        image_path_back = 'above_thearrow_1550 (1).png'
        self.button_icon_back = tkinter.PhotoImage(file=image_path_back)

        # Create a button with the resized image
        self.go_back_but = tkinter.Button(self.frame, text="",command = home, image=self.button_icon_back, compound=tkinter.LEFT, background = '#F0F8FF')
        self.go_back_but.place(x = 650, y = 10)

        #button to make fields appear
        # Load an image for the button icon
        image_path = 'Next_arrow_1559 (1).png'
        self.button_icon = tkinter.PhotoImage(file=image_path)

        # Create a button with the resized image
        self.go_ahead = tkinter.Button(self.frame, text="",command = self.migration_fields, image=self.button_icon, compound=tkinter.LEFT, background = '#F0F8FF')
        self.go_ahead.place(x = 700, y = 10)
        self.go_ahead.config(state = 'disabled')

        #button to upload the .xlsx file
        self.upload_button = customizing.Button(
            frame = self.frame, 
            text = "Upload .xlsx file", 
            command = lambda: (browse_file_xlsx (entry_path=self.entry_path), read_file_xlsx (entry_path_str=self.entry_path.entry.get()), self.sheet_checkboxes ()),
            x = 10,
            y = 10
            )

        self.sheet_1 = customizing.Sheet (frame = frame, tab = self.tab)
        self.sheet_2 = customizing.Sheet (frame = frame, tab = self.tab) 
        self.sheet_3 = customizing.Sheet (frame = frame, tab = self.tab) 
        self.sheet_4 = customizing.Sheet (frame = frame, tab = self.tab) 
        self.sheet_5 = customizing.Sheet (frame = frame, tab = self.tab) 
        self.sheet_6 = customizing.Sheet (frame = frame, tab = self.tab) 
        self.sheet_7 = customizing.Sheet (frame = frame, tab = self.tab) 
        self.sheet_8 = customizing.Sheet (frame = frame, tab = self.tab) 
        self.sheet_9 = customizing.Sheet (frame = frame, tab = self.tab) 
        self.sheet_10 = customizing.Sheet (frame = frame, tab = self.tab) 
        self.sheet_11 = customizing.Sheet (frame = frame, tab = self.tab) 
        self.sheet_12 = customizing.Sheet (frame = frame, tab = self.tab) 
        self.sheet_13 = customizing.Sheet (frame = frame, tab = self.tab) 
        self.sheet_14 = customizing.Sheet (frame = frame, tab = self.tab) 
        self.sheet_15 = customizing.Sheet (frame = frame, tab = self.tab) 
        self.sheet_16 = customizing.Sheet (frame = frame, tab = self.tab) 
        self.sheet_17 = customizing.Sheet (frame = frame, tab = self.tab) 
        self.sheet_18 = customizing.Sheet (frame = frame, tab = self.tab) 
        self.sheet_19 = customizing.Sheet (frame = frame, tab = self.tab) 
        self.sheet_20 = customizing.Sheet (frame = frame, tab = self.tab) 
        self.sheet_21 = customizing.Sheet (frame = frame, tab = self.tab) 
        self.sheet_22 = customizing.Sheet (frame = frame, tab = self.tab) 
        self.sheet_23 = customizing.Sheet (frame = frame, tab = self.tab) 
        self.sheet_24 = customizing.Sheet (frame = frame, tab = self.tab) 
        self.sheet_25 = customizing.Sheet (frame = frame, tab = self.tab) 
        self.sheet_26 = customizing.Sheet (frame = frame, tab = self.tab) 
        self.sheet_27 = customizing.Sheet (frame = frame, tab = self.tab) 
        self.sheet_28 = customizing.Sheet (frame = frame, tab = self.tab) 
        self.sheet_29 = customizing.Sheet (frame = frame, tab = self.tab) 
        self.sheet_30 = customizing.Sheet (frame = frame, tab = self.tab)

        self.download_input_button = customizing.Button (
            frame = self.frame,
            text = 'Download Input Template',
            width = 22
        )
        self.download_input_button.button.place_forget()

        self.upload_input_button = customizing.Button (
            frame = self.frame,
            text = 'Upload Input Template',
            width = 22
        )
        self.upload_input_button.button.place_forget()

        self.upload_input_entry = customizing.Entry (
            frame = frame
        )
        self.upload_input_entry.entry.place_forget()

        self.mode_frame = tkinter.Frame (self.frame, width = 500, height = 400, background = '#F0F8FF')
        self.mode_frame.place (x = 250, y = 100)

        self.mode = customizing.RadioButton_2 (
            frame = self.mode_frame,
            label_text = 'Analysis Mode',
            text_1 = 'Generic',
            dimension = 15
        )
        self.mode_frame.place_forget()


        self.asset_migration_date_label = customizing.Label (
            frame = self.mode_frame,
            text = 'Migration Date',
            dimension = 11
        )
        self.asset_migration_date_label.label.place_forget()

        self.asset_migration_date = customizing.Entry (
            frame = self.mode_frame,
            width = 15
        )
        self.asset_migration_date.entry.place_forget()

        self.asset_depreciation_label = customizing.Label (
            frame = self.mode_frame,
            text = 'Depreciation',
            dimension = 11
        )
        self.asset_depreciation_label.label.place_forget()

        self.asset_depreciation = customizing.Combobox (
            frame = self.mode_frame,
            values = ['Yearly', 'Monthly']
        )
        self.asset_depreciation.combobox.config(state = 'disabled')
        self.asset_depreciation.combobox.place_forget()

        self.calendar = Calendar(self.mode_frame, selectmode="day", date_pattern="dd/mm/yyyy")


        self.sheet_list = [self.sheet_1, self.sheet_2, self.sheet_3, self.sheet_4, self.sheet_5, self.sheet_6, self.sheet_7, self.sheet_8, self.sheet_9, self.sheet_10, self.sheet_11, self.sheet_12, self.sheet_13, self.sheet_14, self.sheet_15, self.sheet_16, self.sheet_17, self.sheet_18, self.sheet_19, self.sheet_20, self.sheet_21, self.sheet_22, self.sheet_23, self.sheet_24, self.sheet_25, self.sheet_26, self.sheet_27, self.sheet_28, self.sheet_29, self.sheet_30]
        
        migration_forget (self.sheet_list, 'Yes')

        # Configure the window to accept file drops
        mainroot.root.drop_target_register(DND_FILES)
        mainroot.root.dnd_bind("<<Drop>>", self.on_drop)
    
    def on_drop(self, event):
        file_path = event.data.strip("{}")
        self.entry_path.entry.config(state = 'normal')
        if file_path:  # Check if a valid file is dropped
            self.entry_path.entry.delete(0, tkinter.END)  # Clear the entry widget
            self.entry_path.entry.insert(0, file_path)  # Insert the selected file path into the entry widget
            read_file_xlsx (entry_path_str=self.entry_path.entry.get())
            self.sheet_checkboxes ()
        else:
            self.entry_path.entry.delete(0, tkinter.END)  # Clear the entry widget
            self.entry_path.entry.insert(0, "Invalid file. Please try again.") # Insert the selected file path into the entry widget
        self.entry_path.entry.config(state = 'disabled')
    
    def go_back (self, screen: str):
        if screen == 'sheets':
            for a in self.sheet_list:
                a.sheet.checkbox.grid_forget()
            self.entry_path.entry.config(state = 'normal')
            self.entry_path.entry.delete(0, tkinter.END)
            self.go_ahead.config(state = 'disabled')
            self.go_back_but.config(command = home)
        elif screen == 'fields':
            start('migration_file', mainframe.frame)
        elif screen == 'input':
            for b in self.sheet_names_list:
                df = pd.read_excel(self.entry_path.entry.get(), b[0])

                # Get the column technical names
                column_tech_names = df.iloc[3, :].tolist()
                # Get the column names
                column_list = df.iloc[6, :].tolist()
                column_names = []
                for column in range(len(column_list)): #if a specific mode is activated, then some fields will be excluded from analysis
                    if self.mode.variable.get() == 'Customer':
                        if column_tech_names[column] not in customizing.mf_customer_general_data and column_tech_names[column] not in customizing.mf_customer_company_data:
                            column_names.append(column_list[column].split('\n')[0])
                    elif self.mode.variable.get() == 'Supplier':
                        if column_tech_names[column] not in customizing.mf_supplier_general_data:
                            column_names.append(column_list[column].split('\n')[0])
                    elif self.mode.variable.get() == 'FI - Accounts receivable open item' or self.mode.variable.get() == 'FI - Accounts payable open item':
                        if column_tech_names[column] not in customizing.mf_bp_open_items:
                            column_names.append(column_list[column].split('\n')[0])
                    elif self.mode.variable.get() == 'FI - G/L account balance and open/line item':
                        if column_tech_names[column] not in customizing.mf_gl_open_items:
                            column_names.append(column_list[column].split('\n')[0])
                    else:
                        column_names.append(column_list[column].split('\n')[0])

                #only the first 50 fields will be included in the analysis (due to a program heaviness); this is the reason for which for specific mode some fields will be excluded
                maximum_field = len(column_names)
                if len(column_names) > 50:
                    maximum_field = 50
                column = 5
                new_line = 0
                for c in range(maximum_field):
                    if c == 8 or c == 16 or c == 24 or c == 32 or c == 40 or c == 48:
                        new_line += 100
                        column = 5
                    self.sheet_list[b[1]].field_list[c].text_input.place_forget()
                    self.sheet_list[b[1]].field_list[c].label.label.place_forget()
                    self.sheet_list[b[1]].field_list[c].label.label.place(x = column, y = new_line)
                    self.sheet_list[b[1]].field_list[c].radiobutton_1.place(x = column, y = new_line + 20)
                    self.sheet_list[b[1]].field_list[c].radiobutton_2.place(x = column, y = new_line + 40)
                    self.sheet_list[b[1]].field_list[c].radiobutton_3.place(x = column, y = new_line + 60)
                    column += 200
            self.download_input_button.button.place_forget()
            self.upload_input_button.button.place_forget()
            self.go_ahead.config(command = self.migration_input)
            self.go_back_but.config(command = lambda: self.go_back('fields'))
    
    def sheet_checkboxes (self): #this function is called when the file is uploaded
        def mode_command (on_off : str): #a function called when you press on mode radiobutton (if you want to use a mode, the main sheet must be present). Using a mode the mandatory sheets must be in the analysis
            if on_off == 'on':
                if not any(element in self.sheet_names for element in customizing.migration_file_main_sheet):
                    tkinter.messagebox.showerror(title="ERROR", message='The main sheet for this template is missing in the Excel file. Choose "Generic" or check the uploaded file')
                    self.mode.variable.set('Generic')
                    return
                for c in range(maximum_sheet):
                    for d in rows:
                        if self.sheet_names[c] in d:
                            if '(mandatory)' in d:
                                self.sheet_list[c].sheet.variable.set(1)
                                self.sheet_list[c].sheet.checkbox.config (state = 'disabled')
                self.asset_depreciation.combobox.config(state = 'normal')
            else:
                for c in range(maximum_sheet):
                    for d in rows:
                        if self.sheet_names[c] in d:
                            if '(mandatory)' in d:
                                self.sheet_list[c].sheet.checkbox.config (state = 'normal')
                self.asset_migration_date.entry.delete(0, tkinter.END)
                self.asset_depreciation.combobox.config(state = 'disabled')

        def insert_date(event):
            if self.mode.variable.get() == 'Fixed asset':
                self.asset_migration_date.entry.config(state = 'normal')
                self.asset_migration_date.entry.delete(0, tkinter.END)  # Clear the entry
                self.asset_migration_date.entry.insert(0, self.calendar.get_date())
                self.asset_migration_date.entry.config(state = 'disabled')

        migration_forget(self.sheet_list, 'Yes')
            
        xl = pd.ExcelFile(self.entry_path.entry.get())

        df = pd.read_excel(self.entry_path.entry.get(), 'Field List')

        column_list = df.columns.tolist()

        all_rows = df.iloc[:, 1].tolist()
        rows = []
        for k in all_rows:
            if isinstance(k, str):
                rows.append(k)

        # Get the sheet names
        self.sheet_names = xl.sheet_names

        #only the first 30 sheets are considered (due to program heaviness)
        maximum_sheet = len(self.sheet_names)
        if len(self.sheet_names) > 30:
            maximum_sheet = 30

        for c in range(maximum_sheet):
            if self.sheet_names[c] != 'Introduction' and self.sheet_names[c] != 'Field List':
                self.sheet_list[c].sheet.checkbox.place(x = 10, y = 25 + 24 * c)
                self.sheet_list[c].sheet.checkbox.config(text = self.sheet_names[c])
        
            for d in rows:
                if self.sheet_names[c] in d:
                    if '(mandatory)' in d:
                        self.sheet_list[c].sheet.variable.set(1) #if a sheet is set as mandatory in the Field List sheet, so by default the checkbox will be ticked
                    else:
                        self.sheet_list[c].sheet.variable.set(0)
        
        for e in customizing.migration_file_modes:
            if e in column_list[0]:
                if e == 'Fixed asset':
                    self.asset_migration_date_label.label.place (x = 150, y = 80)
                    self.asset_migration_date.entry.place (x = 152, y = 110)
                    self.asset_depreciation_label.label.place (x = 255, y = 80)
                    self.asset_depreciation.combobox.place (x = 257, y = 110)
                    self.calendar.place (x = 150, y = 140)
                    # Binding single-click event to the Calendar widget
                    self.calendar.bind("<<CalendarSelected>>", insert_date)
                    self.asset_depreciation.combobox.set('Yearly')
                    self.asset_migration_date.entry.insert(0, (datetime.today().replace(day=1) - timedelta(days=1)).strftime('%d/%m/%Y'))
                    self.asset_migration_date.entry.config(state = 'disabled')
                self.mode.radiobutton_1.config(command = lambda: mode_command('off'))
                self.mode.radiobutton_2.config(text = e, value = e, command = lambda: mode_command('on'))
                self.mode_frame.place (x = 250, y = 100)
                
        self.mode.variable.set('Generic')
        
        self.go_ahead.config(state = 'normal')
        self.go_back_but.config(command = lambda: self.go_back('sheets'))
        self.entry_path.entry.config(state = 'disabled')
    
    def migration_fields (self):
        migration_forget(self.sheet_list, 'No')
        self.mode_frame.place_forget()
        self.sheet_names_list = []
        
        self.tab.place(x = 220, y = 50, height=730, width=1300)

        for b in range(len(self.sheet_list)):
            if self.sheet_list[b].sheet.variable.get() == 1: #only sheets ticked will be considered in the analysis
                column_color = []
                
                self.tab.add(self.sheet_list[b].main_frame.frame, text = self.sheet_names[b])

                df = pd.read_excel(self.entry_path.entry.get(), self.sheet_names[b])
                # You can add code here to extract cell colors using openpyxl
                wb = load_workbook(self.entry_path.entry.get())
                ws = wb[self.sheet_names[b]]
                row = ws[8]
                for cell in row:
                    # Check cell fill color
                    fill_color = cell.fill.start_color.index if cell.fill.start_color.rgb else None
                    column_color.append((cell.column, fill_color))
                wb.close()
                # Now you have both dataframe (df) and cell color information
                # You can further process them according to your requirements

                # Get the column names
                column_tech_names = df.iloc[3, :].tolist()
                column_list = df.iloc[6, :].tolist()
                column_names = []

                # Get the column details
                column_details = df.iloc[4, :].tolist()

                for col_form in range(len(column_details)):
                    if not isinstance(column_tech_names[col_form], str): #check to interrupt analysis if a column has the row 5 (field technical name) blank
                        tkinter.messagebox.showerror(title="ERROR", message=f'Wrong Format in Sheet: {self.sheet_names[b]}, column: {number_to_letters(col_form + 1)}, row: 5')
                        return
                    if not isinstance(column_details[col_form], str): #check to interrupt analysis if a column has the row 6 (technical information) blank
                        tkinter.messagebox.showerror(title="ERROR", message=f'Wrong Format in Sheet: {self.sheet_names[b]}, column: {number_to_letters(col_form + 1)}, row: 6')
                        return
                    if not isinstance(column_list[col_form], str) or "\n" not in column_list[col_form]: #check to interrupt analysis if a column has the row 8 (field name) blank or the format is wrong
                        tkinter.messagebox.showerror(title="ERROR", message=f'Wrong Format in Sheet: {self.sheet_names[b]}, column: {number_to_letters(col_form + 1)}, row: 8')
                        return
                

                for column in range(len(column_list)): #if a specific mode is activated, then some fields will be excluded from analysis
                    color = ''
                    if self.mode.variable.get() == 'Customer':
                        if column_tech_names[column] not in customizing.mf_customer_general_data and column_tech_names[column] not in customizing.mf_customer_company_data:
                            for col in column_color:
                                if col[0] == column + 1:
                                    color = col[1]
                            column_names.append((column_list[column].split('\n')[0], color))
                    elif self.mode.variable.get() == 'Supplier':
                        if column_tech_names[column] not in customizing.mf_supplier_general_data:
                            for col in column_color:
                                if col[0] == column + 1:
                                    color = col[1]
                            column_names.append((column_list[column].split('\n')[0], color))
                    elif self.mode.variable.get() == 'Customer - extend existing record by new org levels':
                        if column_tech_names[column] not in customizing.mf_customer_extend_company_data and column_tech_names[column] not in customizing.mf_customer_extend_sales_data:
                            for col in column_color:
                                if col[0] == column + 1:
                                    color = col[1]
                            column_names.append((column_list[column].split('\n')[0], color))
                    elif self.mode.variable.get() == 'FI - Accounts receivable open item' or self.mode.variable.get() == 'FI - Accounts payable open item':
                        if column_tech_names[column] not in customizing.mf_bp_open_items:
                            for col in column_color:
                                if col[0] == column + 1:
                                    color = col[1]
                            column_names.append((column_list[column].split('\n')[0], color))
                    elif self.mode.variable.get() == 'FI - G/L account balance and open/line item':
                        if column_tech_names[column] not in customizing.mf_gl_open_items:
                            for col in column_color:
                                if col[0] == column + 1:
                                    color = col[1]
                            column_names.append((column_list[column].split('\n')[0], color))
                    else:
                        for col in column_color:
                            if col[0] == column + 1:
                                color = col[1]
                        column_names.append((column_list[column].split('\n')[0], color))
                
                

                #only the first 50 fields will be included in the analysis (due to a program heaviness); this is the reason for which for specific mode some fields will be excluded
                maximum_field = len(column_names)
                if len(column_names) > 50:
                    maximum_field = 50

                column = 5
                new_line = 0
                for a in range(maximum_field):
                    if a == 8 or a == 16 or a == 24 or a == 32 or a == 40 or a == 48:
                        new_line += 100
                        column = 5
                    label_text = column_names[a][0].replace("*", "").replace("+", "")
                    if len(column_names[a][0]) > 25:
                        label_text = column_names[a][0][:25]
                    self.sheet_list[b].field_list[a].label.label.config(text = label_text)
                    self.sheet_list[b].field_list[a].label_text = column_names[a][0]
                    self.sheet_list[b].field_list[a].label.label.place(x = column, y = new_line)
                    self.sheet_list[b].field_list[a].radiobutton_1.place(x = column, y = new_line + 20)
                    self.sheet_list[b].field_list[a].radiobutton_2.place(x = column, y = new_line + 40)
                    self.sheet_list[b].field_list[a].radiobutton_3.place(x = column, y = new_line + 60)
                    if "*" in column_names[a][0] or column_names[a][1] == 'FF92D050':
                        self.sheet_list[b].field_list[a].variable.set('Mandatory') #to have a default for mandatory fields
                        self.sheet_list[b].field_list[a].radiobutton_1.config (background = '#ADFF2F')
                    elif "+" in column_names[a][0] or column_names[a][1] == 'FFFFFF00':
                        self.sheet_list[b].field_list[a].variable.set('Optional') #to have a default for optional fields
                        self.sheet_list[b].field_list[a].radiobutton_2.config (background = '#FFFF66')
                    else:
                        self.sheet_list[b].field_list[a].variable.set('Not Required')
                    column += 200
                
                self.sh_list.append(self.sheet_names[b])
                self.sheet_names_list.append((self.sheet_names[b], b)) #a list with only the sheets to be considered in the analysis

            self.sheet_list[b].sheet.checkbox.config (state = 'disabled')
        
        if self.mode.variable.get() == 'Fixed asset':
            if 'Cumulative Values' in self.sh_list:
                if self.sh_list.index('Cumulative Values') < self.sh_list.index('Posting Information'):
                    tkinter.messagebox.showerror(title="ERROR", message="'Cumulative Values' sheet must be after 'Posting Information' one")
                    return
            if 'Posted Values' in self.sh_list:
                if self.sh_list.index('Posted Values') < self.sh_list.index('Posting Information'):
                    tkinter.messagebox.showerror(title="ERROR", message="'Posted Values' sheet must be after 'Posting Information' one")
                    return
            if 'Transactions (Transf. dur. FY)' in self.sh_list:
                if self.sh_list.index('Transactions (Transf. dur. FY)') < self.sh_list.index('Posting Information'):
                    tkinter.messagebox.showerror(title="ERROR", message="'Transactions (Transf. dur. FY)' sheet must be after 'Posting Information' one")
                    return
            
                    
        self.file_path = self.entry_path.entry.get()
        
        self.upload_button.button.config(state = 'disabled')
        self.go_back_but.config(command = lambda: self.go_back('fields'))
        self.go_ahead.config(command = self.migration_input)
    
    def upload_input_file (self): #this funtion is called when you want to upload template for input values
        for a in self.sheet_names_list:
            df = pd.read_excel(self.upload_input_entry.entry.get(), a[0], dtype=str)

            column_list = df.iloc[0, :].tolist()

            sheet_present_fields = []
            for j in self.present_fields:
                if j[0] == a[0]:
                    sheet_present_fields.append((j[1], j[3], j[4]))
            for b in sheet_present_fields:
                self.sheet_list[a[1]].field_list[b[1]].text_input.delete(1.0, tkinter.END)

                for d in range(len(column_list)):
                    if column_list[d] == b[2]:
                        rows = df.iloc[1:, d].tolist()
                        for row in rows:
                            if str(row) != 'nan':
                                self.sheet_list[a[1]].field_list[b[1]].text_input.insert(tkinter.END, str(row) + '\n')
    
    def migration_input (self, repeat: str = 'No'):
        if repeat == 'No':
            migration_forget(self.sheet_list, 'No')
        self.present_fields = []
        self.error_list = []

        for a in self.sheet_names_list:
            
            df = pd.read_excel(self.file_path, a[0])

            # Get the column names
            column_tech_names = df.iloc[3, :].tolist()
            column_list = df.iloc[6, :].tolist()
            column_names = []
            for column in range(len(column_list)): #similar to the one done in migration fields, but in this case the column number is appended instead of color
                if self.mode.variable.get() == 'Customer':
                    if column_tech_names[column] not in customizing.mf_customer_general_data and column_tech_names[column] not in customizing.mf_customer_company_data:
                        column_names.append((column_list[column].split('\n')[0], column))
                elif self.mode.variable.get() == 'Supplier':
                    if column_tech_names[column] not in customizing.mf_supplier_general_data:
                        column_names.append((column_list[column].split('\n')[0], column))
                elif self.mode.variable.get() == 'Customer - extend existing record by new org levels':
                    if column_tech_names[column] not in customizing.mf_customer_extend_company_data and column_tech_names[column] not in customizing.mf_customer_extend_sales_data:
                        column_names.append((column_list[column].split('\n')[0], column))
                elif self.mode.variable.get() == 'FI - Accounts receivable open item' or self.mode.variable.get() == 'FI - Accounts payable open item':
                    if column_tech_names[column] not in customizing.mf_bp_open_items:
                        column_names.append((column_list[column].split('\n')[0], column))
                elif self.mode.variable.get() == 'FI - G/L account balance and open/line item':
                    if column_tech_names[column] not in customizing.mf_gl_open_items:
                        column_names.append((column_list[column].split('\n')[0], column))
                else:
                    column_names.append((column_list[column].split('\n')[0], column))

            maximum_field = len(column_names)
            if len(column_names) > 50:
                maximum_field = 50
            
            #for specific modes, there is a check that all the fields excluded from the analysis are blank
            if self.mode.variable.get() == 'Customer' and a[0] == 'General Data' or a[0] == 'Company Data':
                for n in range(len(column_tech_names)):
                    all_rows = df.iloc[:, n].tolist()
                    if column_tech_names[n] in customizing.mf_customer_general_data or column_tech_names[n] in customizing.mf_customer_company_data:
                        for row in range(len(all_rows)):
                            if row >= 7:
                                if str(all_rows[row]) != 'nan' or type(all_rows[row]) != float:
                                    self.error_list.append((a[0], 'W001', row + 2 , column_list[n].split('\n')[0].replace("*", "").replace("+", ""), 'This field is not blank, but is in a column not considered in this analysis'))

            elif self.mode.variable.get() == 'Supplier' and a[0] == 'General Data':
                for n in range(len(column_tech_names)):
                    all_rows = df.iloc[:, n].tolist()
                    if column_tech_names[n] in customizing.mf_supplier_general_data:
                        for row in range(len(all_rows)):
                            if row >= 7:
                                if str(all_rows[row]) != 'nan' or type(all_rows[row]) != float:
                                    self.error_list.append((a[0], 'W001', row + 2 , column_list[n].split('\n')[0].replace("*", "").replace("+", ""), 'This field is not blank, but is in a column not considered in this analysis'))

            elif (self.mode.variable.get() == 'FI - Accounts receivable open item' and a[0] == 'Customer Open Items') or (self.mode.variable.get() == 'FI - Accounts payable open item' and a[0] == 'Vendor Open Items'):
                for n in range(len(column_tech_names)):
                    all_rows = df.iloc[:, n].tolist()
                    if column_tech_names[n] in customizing.mf_bp_open_items:
                        for row in range(len(all_rows)):
                            if row >= 7:
                                if str(all_rows[row]) != 'nan' or type(all_rows[row]) != float:
                                    self.error_list.append((a[0], 'W001', row + 2 , column_list[n].split('\n')[0].replace("*", "").replace("+", ""), 'This field is not blank, but is in a column not considered in this analysis'))

            elif self.mode.variable.get() == 'FI - G/L account balance and open/line item' and a[0] == 'GL Balance':
                for n in range(len(column_tech_names)):
                    all_rows = df.iloc[:, n].tolist()
                    if column_tech_names[n] in customizing.mf_gl_open_items:
                        for row in range(len(all_rows)):
                            if row >= 7:
                                if str(all_rows[row]) != 'nan' or type(all_rows[row]) != float:
                                    self.error_list.append((a[0], 'W001', row + 2 , column_list[n].split('\n')[0].replace("*", "").replace("+", ""), 'This field is not blank, but is in a column not considered in this analysis'))

            #checks for mandatory fields (are they filled?) and for not required fields (are they all blank?)
            c = 0
            for b in range(maximum_field):
                rows = df.iloc[:, column_names[b][1]].tolist()
                if self.sheet_list[a[1]].field_list[b].variable.get() == 'Mandatory':
                    self.present_fields.append((a[0], self.sheet_list[a[1]].field_list[b].label_text, column_names[b][1], b, column_tech_names[column_names[b][1]])) #in this list only fields mandatory and optional will be included
                    if repeat == 'No':
                        self.sheet_list[a[1]].field_list[b].label.label.place(x = 5 + 170*c, y = 10)
                        self.sheet_list[a[1]].field_list[b].text_input.place(x = 5 + 170*c, y = 50)
                    for row in range(len(rows)):
                        if row >= 7:
                            if str(rows[row]) == 'nan' and type(rows[row]) == float:
                                self.error_list.append((a[0], 'E001', row + 2 , self.sheet_list[a[1]].field_list[b].label_text.replace("*", "").replace("+", ""), 'This mandatory field is blank'))
                    c+=1

                elif self.sheet_list[a[1]].field_list[b].variable.get() == 'Optional':
                    self.present_fields.append((a[0], self.sheet_list[a[1]].field_list[b].label_text, column_names[b][1], b, column_tech_names[column_names[b][1]]))
                    if repeat == 'No':
                        self.sheet_list[a[1]].field_list[b].label.label.place(x = 5 + 170*c, y = 10)
                        self.sheet_list[a[1]].field_list[b].text_input.place(x = 5 + 170*c, y = 50)
                    c+=1
                
                else:
                    for row in range(len(rows)):
                        if row >= 7:
                            if str(rows[row]) != 'nan' or type(rows[row]) != float:
                                self.error_list.append((a[0], 'E002', row + 2 , self.sheet_list[a[1]].field_list[b].label_text.replace("*", "").replace("+", ""), 'This not required field is filled'))


        if repeat == 'No':
            self.tab.select(0) #in this way the first tab is selected by default

            self.download_input_button.button.config (command = lambda: create_excel_file (self.sheet_names_list, self.present_fields))
            self.upload_input_button.button.config (command = lambda: (browse_file_xlsx(self.upload_input_entry), read_file_xlsx(self.upload_input_entry.entry.get()), self.upload_input_file()))
            self.download_input_button.button.place(x = 800, y = 10)
            self.upload_input_button.button.place(x = 1000, y = 10)

            self.go_back_but.config(command = lambda: self.go_back('input'))
            self.go_ahead.config(command = self.migration_analysis)
    
    def migration_analysis (self, repeat: str = 'No'):
        def repeat_analysis ():
            self.error_list = []
            read_file_xlsx(self.file_path)
            self.migration_input(repeat='Yes')
            self.migration_analysis(repeat='Yes')
            repeat_label = customizing.Label (
                frame = tree.frame,
                text = 'Analysis updated',
                dimension = 12,
                x = 1200,
                y = 10
            )
            repeat_label.label.after (ms = 3000, func = lambda: repeat_label.label.config (text = ''))

        mode_counter = 0
        #lists useful to track speficic fields and do specific transversal checks
        bp_code = []
        postal_code = []
        country = []
        bp_code_tax = []
        tax_type = []
        tax_number = []
        bank_country = []
        bank_key = []
        bank_acc_number = []
        bank_cont_key = []
        iban = []
        finance_bp = []
        logistic_bp = []
        bp_organ = []
        bp_person = []
        logistic_key_fields = []
        asset_cumulative_values = []
        asset_in_current_year = []
        asset_trans_type = []
        asset_trans_amount = []
        transaction_currency = []
        transaction_amount = []
        company_code_currency = []
        company_code_amount = []
        group_currency = []
        group_amount = []
        for a in self.sheet_names_list: #analysis sheets loop
            key_field_list_1 =[]
            key_field_list_2 =[]
            key_field_list_3 =[]

            sheet_present_field = []

            for j in self.present_fields:
                if j[0] == a[0]:
                    sheet_present_field.append((j[2], j[3])) #it identifies the absolute position and the relative position of the analysis fields in the sheet (considering all the fields or only the relevant fields)
            
            df = pd.read_excel(self.file_path, a[0])

            # Get the column names
            column_list = df.iloc[6, :].tolist()
            column_names = []
            for column in column_list:
                column_names.append(column.split('\n')[0])
            
            # Get the column technical names
            column_tech_names = df.iloc[3, :].tolist()
            
            # Get the column details
            column_details = df.iloc[4, :].tolist()
            column_formats = []
            column_int = []
            column_dec = []
            for col_form in range(len(column_details)):
                column_formats.append(column_details[col_form].split(';')[3]) # column formats
            for col_numb in range(len(column_details)): #correction to be made to template maximum length for some fields
                if column_tech_names[col_numb] in customizing.migration_file_1_max_digits:
                    column_int.append('1')
                elif column_tech_names[col_numb] in customizing.migration_file_2_max_digits:
                    column_int.append('2')
                elif column_tech_names[col_numb] in customizing.migration_file_3_max_digits:
                    column_int.append('3')
                elif column_tech_names[col_numb] in customizing.migration_file_4_max_digits:
                    column_int.append('4')
                elif column_tech_names[col_numb] in customizing.migration_file_5_max_digits:
                    column_int.append('5')
                elif column_tech_names[col_numb] in customizing.migration_file_6_max_digits:
                    column_int.append('6')
                elif column_tech_names[col_numb] in customizing.migration_file_7_max_digits:
                    column_int.append('7')
                elif column_tech_names[col_numb] in customizing.migration_file_8_max_digits:
                    column_int.append('8')
                elif column_tech_names[col_numb] in customizing.migration_file_10_max_digits:
                    column_int.append('10')
                elif column_tech_names[col_numb] in customizing.migration_file_28_max_digits:
                    column_int.append('28')
                else:                    
                    column_int.append(column_details[col_numb].split(';')[1]) # column max integer digits
                column_dec.append(column_details[col_numb].split(';')[2]) # column max decimal digits
            
            # Get key columns number
            column_status = df.iloc[5, :].tolist()
            key_counter = 0
            if self.mode.variable.get() in ['FI - Accounts receivable open item', 'FI - Accounts payable open item']:
                key_counter = 4
            elif self.mode.variable.get() == 'FI - G/L account balance and open/line item':
                key_counter = 4
                if ('GL Balance', 'Ledger Group', 1, 1, 'LDGRP') in self.present_fields:
                    key_counter = 5
            else:
                while column_status[key_counter] == 'Key' or str(column_status[key_counter]) == 'nan':
                    key_counter += 1
                    if key_counter == len(column_status):
                        break

            for b in sheet_present_field: #analysis present fields loop
                rows = df.iloc[:, b[0]].tolist()
                input_content = []
                if repeat == 'No':
                    input_content = self.sheet_list[a[1]].field_list[b[1]].text_input.get("1.0", tkinter.END).split('\n')
                    self.input_fields.append((a[1], b[1], self.sheet_list[a[1]].field_list[b[1]].text_input.get("1.0", tkinter.END).split('\n')))
                else:
                    for h in self.input_fields:
                        if h[0] == a[1] and h[1] == b[1]:
                            input_content = h[2]

                for row in range(len(rows)):
                    if row >= 7 and str(rows[row]) != 'nan': #only for rows with template data and fields not blank
                        #format and length controls
                        if column_formats[b[0]] == 'D': #date
                            if not isinstance(rows[row], datetime): # It recognizes both the SAP custom date format and the Excel date format
                                self.error_list.append((a[0], 'E003', row + 2 , self.sheet_list[a[1]].field_list[b[1]].label_text.replace("*", "").replace("+", ""), f'This date is in a wrong format. Insert it using DD/MM/YYYY -> {str(rows[row])}'))

                        elif column_formats[b[0]] == 'N': #integer number
                            if not isinstance(rows[row], int):
                                self.error_list.append((a[0], 'E004', row + 2 , self.sheet_list[a[1]].field_list[b[1]].label_text.replace("*", "").replace("+", ""), f'This number is in a wrong format. Insert it using NNNN -> {str(rows[row])}'))
                            elif rows[row] > 10**int(column_int[b[0]]):
                                self.error_list.append((a[0], 'E005', row + 2 , self.sheet_list[a[1]].field_list[b[1]].label_text.replace("*", "").replace("+", ""), f'The maximum length of the field is exceeded -> {len(str(rows[row]))} > {int(column_int[b[0]])}'))

                        elif column_formats[b[0]] == 'P': #integer or rational number
                            if not isinstance(rows[row], int) and not isinstance(rows[row], float):
                                self.error_list.append((a[0], 'E004', row + 2 , self.sheet_list[a[1]].field_list[b[1]].label_text.replace("*", "").replace("+", ""), f'This number is in a wrong format. Insert it using NNNN.NN -> {str(rows[row])}'))
                            elif isinstance(rows[row], int):
                                if rows[row] > 10**int(column_int[b[0]]):
                                    self.error_list.append((a[0], 'E005', row + 2 , self.sheet_list[a[1]].field_list[b[1]].label_text.replace("*", "").replace("+", ""), f'The maximum length of the field is exceeded -> {len(str(rows[row]))} > {int(column_int[b[0]])}'))
                                if rows[3] in ['WRBTR', 'DMBTR', 'DMBE2']:
                                    if len(str(rows[row]).split('.')[0].replace("-","")) > 11:
                                        self.error_list.append((a[0], 'W010', row + 2 , self.sheet_list[a[1]].field_list[b[1]].label_text.replace("*", "").replace("+", ""), 'The amount is over 100 billion. Check FLETS t-code to see if the amount field is extended'))
                                    elif str(rows[row]).split('.')[0] == '0':
                                        self.error_list.append((a[0], 'W011', row + 2 , self.sheet_list[a[1]].field_list[b[1]].label_text.replace("*", "").replace("+", ""), 'The amount is zero. Check the value'))
                            else:
                                if len(str(rows[row]).split('.')[0]) > int(column_int[b[0]]):
                                    self.error_list.append((a[0], 'E005', row + 2 , self.sheet_list[a[1]].field_list[b[1]].label_text.replace("*", "").replace("+", ""), f"The maximum length of the field is exceeded (integer) -> {len(str(rows[row]).split('.')[0])} > {int(column_int[b[0]])}"))
                                if len(str(rows[row]).split('.')[1]) > int(column_dec[b[0]]):
                                    self.error_list.append((a[0], 'E005', row + 2 , self.sheet_list[a[1]].field_list[b[1]].label_text.replace("*", "").replace("+", ""), f"The maximum length of the field is exceeded (decimals) -> {len(str(rows[row]).split('.')[1])} > {int(column_dec[b[0]])}"))
                                
                                if rows[3] in ['WRBTR', 'DMBTR', 'DMBE2']:
                                    if len(str(rows[row]).split('.')[0].replace("-","")) > 11:
                                        self.error_list.append((a[0], 'W010', row + 2 , self.sheet_list[a[1]].field_list[b[1]].label_text.replace("*", "").replace("+", ""), 'The amount is over 100 billion. Check FLETS t-code to see if the amount field is extended'))

                        else: #char type
                            if len(str(rows[row])) > int(column_int[b[0]]): #check for maximum length exceeded
                                self.error_list.append((a[0], 'E005', row + 2 , self.sheet_list[a[1]].field_list[b[1]].label_text.replace("*", "").replace("+", ""), 'The maximum length of the field is exceeded'))

                            if rows[3] in customizing.migration_file_space_forbidden_fields and (' ' in str(rows[row]) or '\xa0' in str(rows[row])): #check for blank space present in some fields
                                self.error_list.append((a[0], 'E010', row + 2 , self.sheet_list[a[1]].field_list[b[1]].label_text.replace("*", "").replace("+", ""), 'The space in this field is forbidden'))
                            
                            if rows[3] == 'SMTP_ADDR' and '@' not in rows[row]: #check for "@" present in email fields
                                self.error_list.append((a[0], 'E011', row + 2 , self.sheet_list[a[1]].field_list[b[1]].label_text.replace("*", "").replace("+", ""), 'In this field the "@" is mandatory'))

                            if '\n' in rows[row]:
                                self.error_list.append((a[0], 'E031', row + 2 , self.sheet_list[a[1]].field_list[b[1]].label_text.replace("*", "").replace("+", ""), 'In this field go to the next line is forbidden'))

                            #specific checks for modes
                            if self.mode.variable.get() in ['Customer', 'Customer - extend existing record by new org levels']:
                                if rows[3] == 'PARVW' and str(rows[row]) not in customizing.mf_partner_function_customer:
                                    self.error_list.append((a[0], 'E012', row + 2 , self.sheet_list[a[1]].field_list[b[1]].label_text.replace("*", "").replace("+", ""), 'In this field the standard acceptable values are "AG", "RE", "RG", "WE" and "ZM"'))
                            
                            elif self.mode.variable.get() in ['Supplier', 'Supplier - extend existing record by new org levels']:
                                if rows[3] == 'PARVW' and str(rows[row]) not in customizing.mf_partner_function_supplier:
                                    self.error_list.append((a[0], 'E012', row + 2 , self.sheet_list[a[1]].field_list[b[1]].label_text.replace("*", "").replace("+", ""), 'In this field the standard acceptable values are "LF", "WL", "BA", "RS" and "ZM"'))
                                if (rows[3] == 'WT_SUBJCT' or rows[3] == 'WEBRE') and str(rows[row]) != 'X':
                                    self.error_list.append((a[0], 'E012', row + 2 , self.sheet_list[a[1]].field_list[b[1]].label_text.replace("*", "").replace("+", ""), 'In this field the standard acceptable value is "X"'))

                            elif self.mode.variable.get() == 'FI - Accounts receivable open item' or self.mode.variable.get() == 'FI - Accounts payable open item':
                                if rows[3] == 'ZLSPR'  and str(rows[row]) != 'X':
                                    self.error_list.append((a[0], 'E012', row + 2 , self.sheet_list[a[1]].field_list[b[1]].label_text.replace("*", "").replace("+", ""), 'In this field the standard acceptable value is "X"'))

                        #to track all the key fields information
                        if b[1] < key_counter:
                            key_field_list_1.append(rows[row])

                        if not all(element == '' for element in input_content): #check for input fields (blank is always considered)
                            if str(rows[row]) not in input_content:
                                self.error_list.append((a[0], 'E007', row + 2 , self.sheet_list[a[1]].field_list[b[1]].label_text.replace("*", "").replace("+", ""), f'Field filled with a value not foreseen among input fields -> {str(rows[row])}'))                        

                    if row >= 7:
                        if a[0] == 'General Data':
                            if self.mode.variable.get() == 'Customer':
                                if rows[3] == 'KUNNR':
                                    bp_code.append(str(rows[row]))
                                elif rows[3] in ['NAMORG1', 'NAMORG2', 'NAMORG3', 'NAMORG4'] and str(rows[row]) != 'nan' and str(row) not in bp_organ:
                                    bp_organ.append(row)
                                elif rows[3] in ['TITLE_ACA1', 'NAME_FIRST', 'NAME_LAST', 'NAME_MIDDLE', 'NAME_BIRTH', 'BU_SEXID', 'BIRTHDT', 'BIRTHPL', 'LANG_CORR_PERS'] and str(rows[row]) != 'nan' and str(row) not in bp_person:
                                    bp_person.append(str(row))
                                    if str(row) in bp_organ:
                                        self.error_list.append((a[0], 'E017', row + 2 , '', "Both organization data and person data are filled for the same business partner")) 

                            elif self.mode.variable.get() == 'Supplier':
                                if rows[3] == 'LIFNR':
                                    bp_code.append(str(rows[row]))
                                elif rows[3] in ['NAME_FIRST', 'NAME_LAST', 'NAME3', 'NAME4'] and str(rows[row]) != 'nan' and str(row) not in bp_organ:
                                    bp_organ.append(str(row))
                                elif rows[3] in ['TITLE_ACA1', 'NAME_FIRST_P', 'NAME_LAST_P', 'NAME_MIDDLE', 'NAME_BIRTH', 'BU_SEXID', 'BIRTHDT', 'BIRTHPL', 'LANG_CORR_PERS'] and str(rows[row]) != 'nan' and str(row) not in bp_person:
                                    bp_person.append(str(row))
                                    if str(row) in bp_organ:
                                        self.error_list.append((a[0], 'E017', row + 2 , '', "Both organization data and person data are filled for the same business partner")) 

                            if rows[3] == 'POST_CODE1': 
                                postal_code.append(str(rows[row])) #postal code list (transversal check with country)

                            elif rows[3] == 'COUNTRY':
                                country.append(str(rows[row])) #country list
                        
                        elif a[0] == 'Bank Master' or a[0] == 'Bank Details':
                            if rows[3] == 'BANKS':
                                bank_country.append(str(rows[row]))
                            elif rows[3] == 'BANKL':
                                bank_key.append(str(rows[row]))
                            elif rows[3] == 'BANKN':
                                bank_acc_number.append(str(rows[row]))
                            elif rows[3] == 'IBAN':
                                iban.append(str(rows[row]))
                            elif rows[3] == 'BKONT':
                                bank_cont_key.append(str(rows[row]))
                        
                        elif a[0] == 'Tax Numbers' and (self.mode.variable.get() == 'Customer' or self.mode.variable.get() == 'Supplier'):
                            if rows[3] == 'KUNNR' or rows[3] == 'LIFNR':
                                bp_code_tax.append(str(rows[row]))

                            elif rows[3] == 'TAXTYPE':
                                tax_type.append(str(rows[row]))
                            
                            elif rows[3] == 'TAXNUM':
                                tax_number.append(str(rows[row]))

                        if self.mode.variable.get() == 'Fixed asset':
                            if a[0] == 'Posting Information':
                                if rows[3] == 'AKTIV':
                                    if rows[row] > datetime.strptime(self.asset_migration_date.entry.get(), "%d/%m/%Y"):
                                        self.error_list.append((a[0], 'E024', row + 2 , self.sheet_list[a[1]].field_list[b[1]].label_text.replace("*", "").replace("+", ""), f"The date must be prior to the migration date ({self.asset_migration_date.entry.get()})")) 

                                    elif rows[row].year == datetime.strptime(self.asset_migration_date.entry.get(), "%d/%m/%Y").year and not (datetime.strptime(self.asset_migration_date.entry.get(), "%d/%m/%Y").day == 31 and datetime.strptime(self.asset_migration_date.entry.get(), "%d/%m/%Y").month == 12):
                                        asset_in_current_year.append((df.iloc[row, 0], df.iloc[row, 1], df.iloc[row, 2]))
                                        if 'Transactions (Transf. dur. FY)' not in self.sh_list:
                                            self.error_list.append((a[0], 'E025', row + 2 , '', "'Transactions (Transf. dur. FY)' sheet must be included in the analysis, because this asset is capitalized in the migration year")) 

                            if a[0] == 'Cumulative Values':
                                if rows[3] == 'KANSW':
                                    asset_cumulative_values.append((str(rows[row]), row))
                                elif rows[3] == 'KNAFA':
                                    for asset in asset_cumulative_values:
                                        if row == asset[1]:
                                            if (float(asset[0]) > 0 and rows[row] > 0) or (float(asset[0]) < 0 and rows[row] < 0):
                                                self.error_list.append((a[0], 'E022', row + 2 , self.sheet_list[a[1]].field_list[b[1]].label_text.replace("*", "").replace("+", ""), "Cumulative Acq. Values and Acc. Ordinary Depreciation must have different sign")) 
                                            else:
                                                if abs(rows[row]) > abs(float(asset[0])):
                                                    self.error_list.append((a[0], 'E023', row + 2 , self.sheet_list[a[1]].field_list[b[1]].label_text.replace("*", "").replace("+", ""), "Acc. Ordinary Depreciation can not be greater than Cumulative Acq. Values (in absolute value)"))                    

                            if a[0] in ['Cumulative Values', 'Posted Values', 'Transactions (Transf. dur. FY)']:
                                if rows[3] == 'GJAHR':
                                    if rows[row] != datetime.strptime(self.asset_migration_date.entry.get(), "%d/%m/%Y").year:
                                        self.error_list.append((a[0], 'E029', row + 2 , self.sheet_list[a[1]].field_list[b[1]].label_text.replace("*", "").replace("+", ""), "Current fiscal year must be aligned to migration year"))                    

                            if a[0] == 'Transactions (Transf. dur. FY)':
                                if rows [3]== 'ASSETTRTYP':
                                    asset_trans_type.append(str(rows[row]))
                                elif rows[3] == 'AMOUNT':
                                    asset_trans_amount.append (str(rows[row]))
                                elif rows[3] == 'VALUEDATE' and str(rows[row]) != 'nan':
                                    if rows[row] > datetime.strptime(self.asset_migration_date.entry.get(), "%d/%m/%Y"):
                                        self.error_list.append((a[0], 'E024', row + 2 , self.sheet_list[a[1]].field_list[b[1]].label_text.replace("*", "").replace("+", ""), "The date must be prior to the migration date"))
                                    if rows[row].year != datetime.strptime(self.asset_migration_date.entry.get(), "%d/%m/%Y").year:
                                        self.error_list.append((a[0], 'E030', row + 2 , self.sheet_list[a[1]].field_list[b[1]].label_text.replace("*", "").replace("+", ""), "The Reference Date must be in the current year"))  

                        elif self.mode.variable.get() in ['FI - Accounts receivable open item', 'FI - Accounts payable open item', 'FI - G/L account balance and open/line item']:
                            if rows[3] == 'WAERS':
                                transaction_currency.append(str(rows[row]))
                            elif rows[3] == 'WRBTR':
                                transaction_amount.append(str(rows[row]))
                            elif rows[3] == 'HWAER':
                                company_code_currency.append(str(rows[row]))
                            elif rows[3] == 'DMBTR':
                                company_code_amount.append(str(rows[row]))
                            elif rows[3] == 'HWAE2':
                                group_currency.append(str(rows[row]))
                            elif rows[3] == 'DMBE2':
                                group_amount.append(str(rows[row]))

            if a[0] == 'General Data':
                if postal_code != []:
                    for w in range(len(country)):
                        if country[w] in customizing.mf_postal_code_country: #postal code checks
                            error_postal_code = customizing.postal_code_check(postal_code[w], country[w])
                            if error_postal_code != '':
                                self.error_list.append((a[0], 'W002', w + 9 , 'Postal Code', f'The postal code is mandatory by standard for {country[w]}; check OY17 t-code. ' + error_postal_code))
                        else:
                            self.error_list.append((a[0], 'W003', w + 9 , 'Postal Code', f'The postal code for {country[w]} is not analyzed in this program; check OY17 t-code'))
            
            elif a[0] == 'Tax Numbers':
                if self.mode.variable.get() == 'Customer' or self.mode.variable.get() == 'Supplier':
                    for w in range(len(bp_code)):
                        for x in range(len(bp_code_tax)):
                            if bp_code[w] == bp_code_tax[x]: #vat checks
                                if country[w] in customizing.mf_vat_country:
                                    error_vat = customizing.vat_check(tax_type[x], tax_number[x], country[w])
                                    if error_vat != '':
                                        self.error_list.append((a[0], 'W004', x + 9 , 'Tax Number', error_vat))
                                else:
                                    self.error_list.append((a[0], 'W005', x + 9 , 'Tax Number', f'The VAT number for {country[w]} is not analyzed in this program'))

            elif a[0] == 'Bank Master' or a[0] == 'Bank Details':
                for y in range(len(bank_country)):
                    if bank_country[y] in customizing.mf_bank_country:
                        if a[0] == 'Bank Master': #bank data checks
                            error_bank = customizing.bank_check(a[0], bank_country[y], bank_key[y])
                        else:
                            error_bank = customizing.bank_check(a[0], bank_country[y], bank_key[y], bank_acc_number[y], bank_cont_key[y], iban[y])
                        if error_bank != '':
                            self.error_list.append((a[0], 'W006', y + 9 , '', error_bank + '; check 0Y17 t-code'))
                    else:
                        self.error_list.append((a[0], 'W007', y + 9 , '', f'The bank data for {bank_country[y]} is not analyzed in this program; check OY17 t-code'))

            if self.mode.variable.get() == 'Fixed asset':
                if asset_trans_type != []:
                    for no, tr_type in enumerate (asset_trans_type):
                        if str(tr_type) == '100' and float(asset_trans_amount[no]) < 0:
                            self.error_list.append((a[0], 'W015', no + 9 , '', 'With "100" as Asset Transaction Type the sign should be positive'))
                        if (str(tr_type) == '200' or str(tr_type) == '210') and float(asset_trans_amount[no]) > 0:
                            self.error_list.append((a[0], 'W016', no + 9 , '', f'With "{tr_type}" as Asset Transaction Type the sign should be negative'))

            if self.mode.variable.get() in ['FI - Accounts receivable open item', 'FI - Accounts payable open item', 'FI - G/L account balance and open/line item']:
                for num, t_curr in enumerate(transaction_currency):
                    if company_code_currency != []:
                        if t_curr == company_code_currency[num]:
                            if transaction_amount[num] != company_code_amount[num]:
                                self.error_list.append((a[0], 'W012', num + 9 , 'Amount', f'Transaction Amount and Company Code Amount should be equal, since the currency is {t_curr} in both of them'))
                    if group_currency != []:
                        if t_curr == group_currency[num]:
                            if transaction_amount[num] != group_amount[num]:
                                self.error_list.append((a[0], 'W013', num + 9 , 'Amount', f'Transaction Amount and Group Amount should be equal, since the currency is {t_curr} in both of them'))
                    if company_code_currency != [] and group_currency != []:
                        if company_code_currency[num] == group_currency[num]:
                            if company_code_amount[num] != group_amount[num]:
                                self.error_list.append((a[0], 'W014', num + 9 , 'Amount', f'Company Code Amount and Group Amount should be equal, since the currency is {company_code_currency[num]} in both of them'))

                    if (float(transaction_amount[num]) > 0 and (float(company_code_amount[num]) < 0 or float(group_amount[num]) < 0)) or (float(transaction_amount[num]) < 0 and (float(company_code_amount[num]) > 0 or float(group_amount[num]) > 0)):
                        self.error_list.append((a[0], 'E032', num + 9 , 'Amount', 'Transaction Amount, Company Code Amount and Group Amount can not have different sign'))

            for c in range(int(len(key_field_list_1)/key_counter)): #check to verify that key fields are not doubled in the same sheet
                counter = 0
                for d in range(key_counter):
                    key_field_list_2.append(key_field_list_1[counter + c]) #to sort the key field values
                    counter += int(len(key_field_list_1)/key_counter)
                if key_field_list_2 in key_field_list_3: #check if key field value is already present in the sheet
                    self.error_list.append((a[0], 'E006', c + 9 , '', f'These key field values are already present in the sheet -> {key_field_list_2}'))
                key_field_list_3.append(key_field_list_2)
                key_field_list_2 = []
            
            if a[0] in ['Purchasing Organization Data', 'Sales Data']:
                logistic_key_fields = key_field_list_3
            
            #this check is done only if a specific mode is chosen (it verify that in other sheets, the key values of the main sheet are used)
            main_key_field = []
            main_key_fields_list = []
            if self.mode.variable.get() in customizing.migration_file_modes:
                if a[0] in customizing.migration_file_main_sheet: 
                    self.mode_key_fields = key_field_list_3 #to track all the key field values for the main sheet
                    mode_counter = len(key_field_list_3[0])
                else:
                    for k in range(len(key_field_list_3)):
                        main_key_field = key_field_list_3[k][:mode_counter]
                        if main_key_field not in self.mode_key_fields:
                            self.error_list.append((a[0], 'E008', k + 9 , '', f'These key field values are not present in the main sheet -> {main_key_field}'))                         
                        main_key_fields_list.append(main_key_field)

                    if a[0] in customizing.migration_file_secondary_sheets: #for secondary sheets, all the key values of the main sheets must be present
                        for n in range(len(self.mode_key_fields)):
                            if self.mode_key_fields[n] not in main_key_fields_list:
                                self.error_list.append((a[0], 'E009', '' , '', f'The {self.mode_key_fields[n]} key field values are not in this sheet'))                         

                    bp_counter = 2
                    if self.mode.variable.get() in ['Customer', 'Supplier']:
                        bp_counter = 1
                    if a[0] == 'BP Roles': #to extract only BP open in finance/logistic
                        for bp in key_field_list_3:
                            if bp[bp_counter] in ['FLVN00', 'FLCU00'] and bp[:bp_counter] not in finance_bp:
                                finance_bp.append(bp[:bp_counter])
                            elif bp[bp_counter] in ['FLVN01', 'FLCU01'] and bp[:bp_counter] not in logistic_bp:
                                logistic_bp.append(bp[:bp_counter])

                    if a[0] in customizing.migration_file_finance_bp_sheets:
                        bp_finance = []
                        for bp_fin in key_field_list_3:
                            bp_finance.append(bp_fin[:bp_counter])
                            if bp_fin[:bp_counter] not in finance_bp:
                                self.error_list.append((a[0], 'E013', '' , '', f'The {bp_fin[:bp_counter]} business partner is not open in Finance in the "BP Roles" sheet'))                         

                        if a[0] in customizing.migration_file_finance_bp_complete_sheets:
                            for bp_fin_compl in finance_bp:
                                if bp_fin_compl not in bp_finance:
                                    self.error_list.append((a[0], 'E014', '' , '', f'The {bp_fin_compl} business partner is open in Finance in the "BP Roles" sheet, but it is not present in this sheet'))                         

                    if a[0] in customizing.migration_file_logistic_bp_sheets:
                        bp_logistic = []
                        for bp_log in key_field_list_3:
                            bp_logistic.append(bp_log[:bp_counter])
                            if bp_log[:bp_counter] not in logistic_bp:
                                self.error_list.append((a[0], 'E015', '' , '', f'The {bp_log[:bp_counter]} business partner is not open in Purchasing/Sales in the "BP Roles" sheet'))                         

                        if a[0] in customizing.migration_file_logistic_bp_complete_sheets:
                            for bp_log_compl in logistic_bp:
                                if bp_log_compl not in bp_logistic:
                                    self.error_list.append((a[0], 'E016', '' , '', f'The {bp_log_compl} business partner is open in Purchasing/Sales in the "BP Roles" sheet, but it is not present in this sheet'))                         

                    key_fields_list = []
                    key_fields_list_trans = []
                    if a[0] in customizing.migration_file_purchasing_bp_dependent_sheets:
                        if logistic_key_fields == []:
                            tkinter.messagebox.showerror(title="ERROR", message=f'The Purchasing Organization Data is put after the Partner Functions one or is completely empty. Check the file')
                            return
                        for k in range(len(key_field_list_3)):
                            key_field = key_field_list_3[k][:(bp_counter + 1)]
                            if key_field not in logistic_key_fields:
                                self.error_list.append((a[0], 'E018', k + 9 , '', f'These key field values are not present in the Purchasing Organization Data sheet -> {key_field}'))                         
                            key_fields_list.append(key_field)

                        for n in range(len(logistic_key_fields)):
                            if logistic_key_fields[n] not in key_fields_list:
                                self.error_list.append((a[0], 'E019', '' , '', f'The {logistic_key_fields[n]} key field values are present in the Purchasing Organization Data sheet but not in this one'))                         

                    if a[0] in customizing.migration_file_sales_bp_dependent_sheets:
                        if logistic_key_fields == []:
                            tkinter.messagebox.showerror(title="ERROR", message=f'The Sales Data is put after the Sales Partner one or is completely empty. Check the file')
                            return
                        for k in range(len(key_field_list_3)):
                            key_field = key_field_list_3[k][:(bp_counter + 3)]
                            if key_field not in logistic_key_fields:
                                self.error_list.append((a[0], 'E020', k + 9 , '', f'These key field values are not present in the Sales Data sheet -> {key_field}'))                         
                            key_fields_list.append(key_field)

                        for n in range(len(logistic_key_fields)):
                            if logistic_key_fields[n] not in key_fields_list:
                                self.error_list.append((a[0], 'E021', '' , '', f'The {logistic_key_fields[n]} key field values are present in the Sales Data sheet but not in this one'))                    

                    if self.mode.variable.get() == 'Fixed asset':
                        if a[0] == 'Cumulative Values':
                            for key in range(len(key_field_list_3)):
                                key_field = tuple(key_field_list_3[key][:3])
                                if key_field in asset_in_current_year:
                                    self.error_list.append((a[0], 'E026', key + 9 , '', f"This asset must not be tracked in 'Cumulative Values' sheet, but only in 'Transactions (Transf. dur. FY)' one -> {key_field}"))                         
                                key_fields_list.append(key_field)
                            for ass in self.mode_key_fields:
                                if tuple(ass) not in asset_in_current_year and tuple(ass) not in key_fields_list:
                                    self.error_list.append((a[0], 'E028', '' , '', f"The {ass} asset is not present in this sheet"))                         

                        if a[0] == 'Transactions (Transf. dur. FY)':
                            for key in range(len(key_field_list_3)):
                                key_field_trans = tuple(key_field_list_3[key][:3])
                                key_fields_list_trans.append(key_field_trans)
                            for curr in asset_in_current_year:
                                if curr not in key_fields_list_trans: 
                                    self.error_list.append((a[0], 'E027', '' , '', f"The {curr} asset acquisition must be tracked in this sheet, since it's an asset capitalized in the migration year"))                         

                        if a[0] == 'Posted Values':
                            if self.asset_depreciation.combobox.get() == 'Monthly' and not (datetime.strptime(self.asset_migration_date.entry.get(), "%d/%m/%Y").day == 31 and datetime.strptime(self.asset_migration_date.entry.get(), "%d/%m/%Y").month == 12):
                                if key_field_list_3 == []:
                                    self.error_list.append((a[0], 'W008', '' , '', "For monthly depreciation and migration during the year, it needs to use this sheet to track depreciations in the current year"))                         
                            else:
                                if key_field_list_3 != []:
                                    self.error_list.append((a[0], 'W009', '' , '', "It needs to use this sheet only to track depreciations in the current year, for monthly depreciation and migration during the year"))                         

        for widget in self.frame.winfo_children():
            widget.destroy()   

        lst = [('SHEET', 'MESSAGE ID', 'ROW', 'FIELD', 'ERROR MESSAGE')]
        errors_number = 0
        warnings_number = 0
        for a in range(len(self.error_list)):
            lst.append(self.error_list[a])
            if 'E' in self.error_list[a][1]:
                errors_number += 1
            elif 'W' in self.error_list[a][1]:
                warnings_number += 1

        tree = customizing.Treeview (
            frame = self.frame,
            col_text = ['', '', '', '', ''],
            width_list = [120, 30, 30, 120, 540],
            lst = lst
        )

        image_error_path = 'red_cross.png'
        image_warning_path = 'yellow_triangle.png'

        # Open the image using Pillow
        image_error_img = Image.open(image_error_path)
        image_warning_img = Image.open(image_warning_path)
        # Convert the image to a format Tkinter supports
        image_error = ImageTk.PhotoImage(image_error_img)
        image_warning = ImageTk.PhotoImage(image_warning_img)

        label_error = tkinter.Label(tree.frame, image=image_error, text = f'Errors: {errors_number}', compound=tkinter.LEFT, font = ('Calibri', 14, 'bold'), background = '#F0F8FF')
        label_error.place (x = 400, y = 10)

        label_warning = tkinter.Label(tree.frame, image=image_warning, text = f'Warnings: {warnings_number}', compound=tkinter.LEFT, font = ('Calibri', 14, 'bold'), background = '#F0F8FF')
        label_warning.place (x = 600, y = 10)

        # Keep a reference to the image to prevent it from being garbage collected
        label_error.image = image_error
        label_warning.image = image_warning

        

        repeat_analysis_button = customizing.Button (
            frame = tree.frame,
            text = 'Repeat Analysis',
            command = repeat_analysis,
            x = 1000,
            y = 10
            )

mainroot = customizing.Root (root_title = 'SAP HELPER')

mainframe = customizing.Frame (
            root = mainroot.root,
            pack_or_grid = 'P'
        )

main_menubar = customizing.MenuBar (
    root = mainroot.root,
    first_label = 'File',
    second_label = 'Go To',
    third_label = 'Help'
)

main_menubar.main_menu_1.add_command (label = 'Quit', command = lambda: quit_window (mainroot.root), font = ('Calibri', 12))
main_menubar.main_menu_2.add_command (label = 'Home', command = home, font = ('Calibri', 12))
main_menubar.main_menu_2.add_command (label = 'Ebs MT940', command = lambda: start('ebs_mt940', mainframe.frame), font = ('Calibri', 12))
main_menubar.main_menu_2.add_command (label = 'IBAN', command = lambda: start('iban', mainframe.frame), font = ('Calibri', 12))
main_menubar.main_menu_2.add_command (label = 'Migration File', command = lambda: start('migration_file', mainframe.frame), font = ('Calibri', 12))
main_menubar.main_menu_3.add_command (label = 'EBS MT940 info', command = lambda: info('EBS MT940 info'), font = ('Calibri', 12))
main_menubar.main_menu_3.add_command (label = 'IBAN info', command = lambda: info('IBAN info'), font = ('Calibri', 12))
main_menubar.main_menu_3.add_command (label = 'Migration File info', command = lambda: info('Migration File info'), font = ('Calibri', 12))


main = MainRoot (root = mainroot.root, frame = mainframe.frame)



mainroot.root.mainloop()