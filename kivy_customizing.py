from kivy.app import App
from kivy.uix.label import Label
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.widget import Widget
from kivy.graphics import Color, Line
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.core.window import Window
from kivy.uix.popup import Popup
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.uix.dropdown import DropDown
from kivy.uix.scrollview import ScrollView
from kivy.uix.gridlayout import GridLayout
from datetime import datetime
from kivymd.app import MDApp
import ctypes
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font

#dictionary with all the known bank transaction types
ebs_mt940_dict = {
    'FCHG': 'Charges and other expenses',
    'FCHK': 'Cheques',
    'FINT': 'Interest',
    'FRTI': 'Returned item',
    'NBOE': 'Bill of exchange',
    'NCHG': 'Charges and other expenses',
    'NCHK': 'Cheques',
    'NCOL': 'Collections',
    'NCOM': 'Commisions',
    'NDCR': 'Documentary credit',
    'NDDT': 'Direct Debit Item',
    'NDIV': 'Dividends-Warrants',
    'NECK': 'Eurocheques',
    'NEQA': 'Equivalent amount',
    'NFEX': 'Foreign exchange',
    'NINT': 'Interest',
    'NLDP': 'Loan deposit',
    'NMSC': 'Miscellaneous',
    'NRTI': 'Returned item',
    'NSEC': 'Securities',
    'NSTO': 'Standing order',
    'NTCK': 'Travellers cheques',
    'NTRF': 'Transfer',
    'S100': 'SWIFT message 100',
    'S101': 'SWIFT message 101',
    'S103': 'SWIFT message 103',
    'S190': 'SWIFT message 190',
    'S191': 'SWIFT message 191',
    'S200': 'SWIFT message 200',
    'S201': 'SWIFT message 201',
    'S202': 'SWIFT message 202',
    'S203': 'SWIFT message 203',
    'S205': 'SWIFT message 205',
    'S300': 'SWIFT message 300',
    'S320': 'SWIFT message 320',
    'S400': 'SWIFT message 400',
    'S554': 'SWIFT message 554',
    'S556': 'SWIFT message 556'
}

def truncate_string(value, max_length=30):
    if len(value) > max_length:
        return value[:max_length - 3] + "..."  # Truncate and append "..."
    return value

class KivyButton ():
    def __init__ (
        self,
        layout: FloatLayout,
        text: str = '',
        font_size: str = '24sp',
        size_hint = (0.20, 0.08),
        bg_color = (0.54, 0.81, 0.94, 1),
        color: str = (0, 0, 0, 1),
        width: int = 64,
        height: int = 64,
        background_normal = '',
        background_down = '',
        background_disabled_normal = '',
        x: float = 0,
        y: float = 0
    ):
        if background_normal == '':
            # Create a button
            self.button = Button(
                text=text,
                font_size=font_size,  # Set font size (in scalable pixels)
                size_hint=size_hint,  
                pos_hint={"center_x": x, "center_y": y},
                background_color=bg_color,  # RGBA color
                background_normal='',  # Remove default background image
                color=color
            )
        else:
            # Create a button
            self.button = Button(
                size_hint=(None, None),
                width = width,
                height = height,
                pos_hint={'center_x': x, 'center_y': y},  # Adjust position as needed
                background_normal = background_normal,
                background_down = background_down,
                background_disabled_normal = background_disabled_normal
            )
        layout.add_widget(self.button)

class KivyTable:
    def __init__(self, scroll_view: ScrollView, layout: FloatLayout, headers: list, rows: list, widths: list):
        self.scroll_view = scroll_view
        self.layout = layout
        self.headers = headers
        self.rows = rows
        self.widths = widths

        row_height = 40  # Fixed row height
        total_height = len(rows) * row_height

        # Create a scrollable content area
        self.scroll_content = FloatLayout(size_hint_y=None)
        self.scroll_content.height = total_height

        # Create headers
        y_position = total_height
        x_position = 0
        for i, header in enumerate(headers):
            header_label = Label(
                text=header,
                size_hint=(None, None),
                size=(widths[i], row_height),
                pos=(x_position, 870),
                color=(0, 0, 0, 1),
                halign='left',
                valign='middle',
                text_size=(widths[i], row_height),
                padding=(10, 0),
                bold=True
            )
            self.layout.add_widget(header_label)
            x_position += widths[i]

        # Create rows
        y_position -= row_height  # Start below headers
        for row in rows:
            x_position = 0
            for i, cell in enumerate(row):
                cell_label = Label(
                    text=str(cell),
                    size_hint=(None, None),
                    size=(widths[i], row_height),
                    pos=(x_position, y_position),
                    color=(0, 0, 0, 1),
                    halign='left',
                    valign='middle',
                    text_size=(3000, row_height),
                    width = 3000,
                    padding=(10, 0)
                )
                self.scroll_content.add_widget(cell_label)
                x_position += widths[i]
            y_position -= row_height

        # Add scrollable content to the scroll view
        self.scroll_view.add_widget(self.scroll_content)

def export_to_excel(layout: FloatLayout, table: KivyTable, file_path: str, file_name: str, ebs: str = "NO"):
    # Extract data from GridLayout
    data = []
    row = []
    if ebs == 'NO':
        data = [table.headers]
    count = 0

    for child in table.rows:  # Reversed because GridLayout adds widgets in reverse order
        data.append(child)

    # Create a Pandas DataFrame
    df = pd.DataFrame(data)

    current_datetime_str = datetime.now().strftime("%Y-%m-%d %H.%M.%S")

    father_path = file_path.rsplit("\\", 1)[0]

    # Save the DataFrame to an Excel file
    file_path = father_path + "\\" + file_name + f" - {current_datetime_str}.xlsx"
    df.to_excel(file_path, index=False, header=False)


    # Load the workbook to adjust column widths
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Apply bold font to the first and fourth rows
    bold_font = Font(bold=True)
    
    for row_num, row_cells in enumerate(sheet.iter_rows(), start=1):
        if ebs == "YES" and (row_num == 1 or row_num == 4):  # Make first and fourth rows bold
            for cell in row_cells:
                cell.font = bold_font
        if row_num == 1: # Make first row bold
            for cell in row_cells:
                cell.font = bold_font

    
    # Auto-adjust the column widths
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)  # Adjust for padding
        sheet.column_dimensions[column].width = adjusted_width

    # Save the workbook with adjusted column widths
    workbook.save(file_path)

    save_label = Label (
        text=f'The file was saved in {father_path}',
        font_size='16sp',
        color=(0, 0.5, 0, 1),  # color for text
        pos_hint={'x': 0.16, 'y': 0.02},
        size_hint=(0.1, 0.04)
        )
    
    layout.add_widget(save_label)