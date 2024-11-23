import tkinter
from tkinter import ttk
import tkinter.messagebox
import pyperclip
import re
from tkinterdnd2 import DND_FILES, TkinterDnD
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font

#dictionary with all the known bank transaction types
ebs_mt940_dict = {
    'FCHG': 'Charges and other expenses',
    'FCHK': 'Cheques',
    'FINT': 'Interest',
    'FRTI': 'Returned item',
    'NBOE': 'Bill of exchange',
    'NCHG': 'Charges and other expenses',
    'NCHK': 'Cheques',
    'NCOL': 'Collections',
    'NCOM': 'Commisions',
    'NDCR': 'Documentary credit',
    'NDDT': 'Direct Debit Item',
    'NDIV': 'Dividends-Warrants',
    'NECK': 'Eurocheques',
    'NEQA': 'Equivalent amount',
    'NFEX': 'Foreign exchange',
    'NINT': 'Interest',
    'NLDP': 'Loan deposit',
    'NMSC': 'Miscellaneous',
    'NRTI': 'Returned item',
    'NSEC': 'Securities',
    'NSTO': 'Standing order',
    'NTCK': 'Travellers cheques',
    'NTRF': 'Transfer',
    'S100': 'SWIFT message 100',
    'S101': 'SWIFT message 101',
    'S103': 'SWIFT message 103',
    'S190': 'SWIFT message 190',
    'S191': 'SWIFT message 191',
    'S200': 'SWIFT message 200',
    'S201': 'SWIFT message 201',
    'S202': 'SWIFT message 202',
    'S203': 'SWIFT message 203',
    'S205': 'SWIFT message 205',
    'S300': 'SWIFT message 300',
    'S320': 'SWIFT message 320',
    'S400': 'SWIFT message 400',
    'S554': 'SWIFT message 554',
    'S556': 'SWIFT message 556',

}

#specific modes to implement additional checks for these templates
migration_file_modes = ['Fixed asset', 'Customer', 'Customer - extend existing record by new org levels', 'Supplier', 'Supplier - extend existing record by new org levels', 'FI - Accounts receivable open item', 'FI - Accounts payable open item', 'FI - G/L account balance and open/line item']

migration_file_main_sheet = ['Master Details', 'General Data', 'Customer Open Items', 'Vendor Open Items', 'GL Balance']

#a list of sheets for which is mandatory to have all the key value of the main sheet
migration_file_secondary_sheets = ['Posting Information', 'Time-Dependent Data', 'Depreciation Areas', #ASSET
                                   'BP Roles'] #CUSTOMER/VENDOR

migration_file_finance_bp_sheets = ['Company Data', 'Withholding Tax Data']

migration_file_finance_bp_complete_sheets = ['Company Data'] #sheets for which all the BP open in finance, need to be filled

migration_file_logistic_bp_sheets = ['Sales Data', 'Output Tax', 'Purchasing Organization Data']

migration_file_logistic_bp_complete_sheets = ['Sales Data', 'Output Tax', 'Purchasing Organization Data'] #sheets for which all the BP open in sales/purchasing, need to be filled

migration_file_purchasing_bp_dependent_sheets = ['Partner Functions']

migration_file_sales_bp_dependent_sheets = ['Sales Partner']

#a list of technical name fields for which is forbidden to have spaces
migration_file_space_forbidden_fields = ['BUKRS', 'ANLN2', 'ANLKL', 'GSBER', 'KOSTL', 'WERKS', 'AFABE', 'ASSETTRTYP', 'CURRENCY', #ASSET
                                         'KUNNR', 'BU_GROUP', 'VBUND', 'BPEXT', 'COUNTRY', 'REGION', 'LANGU_CORR', 'SMTP_ADDR', #CUSTOMER - GENERAL DATA
                                         'BP_ROLE', 'MAHNA', 'ZTERM', 'ZWELS_01', 'ZWELS_02', 'ZWELS_03', 'ZWELS_04', 'HBKID', 'AKONT', 'WITHT', 'WT_WITHCD', #CUSTOMER - BP ROLES/COMPANY DATA/WHT
                                         'VKORG', 'VTWEG', 'SPART', 'KDGRP', 'BZIRK', 'VKBUR', 'WAERS', 'KONDA', 'KALKS', 'LPRIO', 'VSBED', 'INCO1', 'KTGRD', 'PARVW', 'KUNN2', 'ALAND', 'TATYP', 'TAXKD', #CUSTOMER - SALES DATA/SALES PARTNER/OUTPUT TAX
                                         'BANKS', 'BANKL', 'BANKN', 'IBAN', 'BKONT', 'TAXTYPE', 'TAXNUM', #CUSTOMER - BANK DATA/TAX NUMBER 
                                         'LIFNR', 'FRGRP', 'ZTERM1', 'REPRF', 'WT_SUBJCT', #VENDOR - COMPANY DATA
                                         'EKORG', 'EKGRP', 'WEBRE', 'BSTAE', 'LIFN2', #VENDOR - PURCHASING ORGANIZATION DATA/PARTNER FUNCTION
                                         'PROVZ', 'SWIFT', #BANK
                                         'GKONT', 'UMSKZ', 'HWAER', 'HWAE2', 'HWAE3', 'MWSKZ', 'ZLSCH', 'ZLSPR', 'PRCTR', 'FKBER', 'PSPNR', 'WT_TYPE', 'WT_CODE', #CUSTOMER/VENDOR OI
                                         'RASSC', 'COPA_PRCTR'] #GL OI

#fields not to be considered in the related sheet for customer template (there is only a check about these fields are blank)
mf_customer_general_data = ['LEGAL_ENTY', 'LEGAL_ORG', 'FOUND_DAT', 'LIQUID_DAT', 'LOCATION_1', 'LOCATION_2', 'LOCATION_3', 'BAHNE', 'BAHNS', 'COUNC', 'CITYC', 'DTAMS', 'DTAWS', 'KNRZA', 'NIELS', 'RPMKR', 'KUKLA', 'HZUOR', 'BRAN1', 'BRAN2', 'BRAN3', 'BRAN4', 'BRAN5', 'KATR1', 'KATR2', 'KATR3', 'KATR4', 'KATR5', 'KATR6', 'KATR7', 'KATR8', 'KATR9', 'KATR10', 'SUFRAMA', 'RG', 'EXP', 'UF', 'RGDATE', 'RIC', 'RNE', 'RNEDATE', 'CNAE', 'LEGALNAT', 'CRTN', 'ICMSTAXPAY', 'INDTYP', 'TDT', 'COMSIZE', 'DECREGPC', 'ECC_NO', 'EXC_REG_NO', 'EXC_RANGE', 'EXC_DIV', 'EXC_COMM', 'EXC_TAX_IND', 'CST_NO', 'LST_NO', 'SERV_TAX_NO', 'PAN_NO', 'PAN_REF_NO', 'BON_AREA_CONF', 'DON_MARK', 'CONSOLIDATE_INVOICE', 'ALLOWANCE_TYPE', 'EINVOICE_MODE', 'J_1KFTBUS', 'J_1KFTIND', 'J_1KFREPRE', 'PH_BIZ_STYLE', 'CITY2', 'HOME_CITY', 'TIME_ZONE', 'LZONE', 'BUILDING', 'ROOM', 'FLOOR', 'CO_NAME', 'HOUSE_NO2', 'STR_SUPPL3', 'LOCATION', 'TXJCD', 'NOTE_TELNR', 'TELNR_LONG_2', 'NOTE_TELNR_2', 'TELNR_LONG_3', 'NOTE_TELNR_3', 'NOTE_MOBILE', 'MOBILE_LONG_2', 'NOTE_MOBILE_2', 'MOBILE_LONG_3', 'NOTE_MOBILE_3', 'NOTE_FAXNR', 'FAXNR_LONG_2', 'NOTE_FAXNR_2', 'FAXNR_LONG_3', 'NOTE_FAXNR_3', 'NOTE_SMTP', 'SMTP_ADDR_2', 'NOTE_SMTP_2', 'SMTP_ADDR_3', 'NOTE_SMTP_3', 'URI_TYP', 'URI_ADDR', 'NOTE_URI', 'SPERR', 'COLLMAN']
mf_customer_company_data = ['TLFNS', 'TLFXS', 'INTAD']

#fields not to be considered in the related sheet for supplier template (there is only a check about these fields are blank)
mf_supplier_general_data = ['LEGAL_ENTY', 'LEGAL_ORG', 'FOUND_DAT', 'LIQUID_DAT', 'LOCATION_1', 'LOCATION_2', 'LOCATION_3', 'DTAMS', 'DTAWS', 'LNRZA', 'ESRNR', 'TERM_LI', 'MIN_COMP', 'COMSIZE', 'DECREGPC', 'CRC_NUM', 'RG', 'EXP', 'UF', 'RGDATE', 'RIC', 'RNE', 'RNEDATE', 'CNAE', 'LEGALNAT', 'CRTN', 'ICMSTAXPAY', 'INDTYP', 'TDT', 'J_1IEXCD', 'J_1IEXRN', 'J_1IEXRG', 'J_1IEXDI', 'J_1IEXCO', 'J_1IVTYP', 'J_1I_CUSTOMS', 'J_1IEXCIVE', 'J_1ISSIST', 'J_1IVENCRE', 'J_1ICSTNO', 'J_1ILSTNO', 'J_1ISERN', 'J_1IPANNO', 'J_1IPANREF', 'J_1IPANVALDT', 'J_1IDEDREF', 'VEN_CLASS', 'J_1KFTBUS', 'J_1KFTIND', 'J_1KFREPRE', 'CATEG', 'STATUS', 'VFNUM', 'VFNID', 'PARTNER_NAME', 'PARTNER_UTR', 'CRN', 'ALLOWANCE_TYPE', 'AU_CARRYING_ENT', 'AU_IND_UNDER_18', 'AU_PAYMENT_NOT_EXCEED_75', 'AU_WHOLLY_INP_TAXED', 'AU_PARTNER_WITHOUT_GAIN', 'AU_NOT_ENTITLED_ABN', 'AU_PAYMENT_EXEMPT', 'AU_PRIVATE_HOBBY', 'AU_DOMESTIC_NATURE', 'SC_CAPITAL', 'SC_CURRENCY', 'CITY2', 'HOME_CITY', 'TIME_ZONE', 'LZONE', 'BUILDING', 'ROOM', 'FLOOR', 'CO_NAME', 'HOUSE_NO2', 'STR_SUPPL3', 'LOCATION', 'TXJCD', 'NOTE_TELNR', 'TELNR_LONG_2', 'NOTE_TELNR_2', 'TELNR_LONG_3', 'NOTE_TELNR_3', 'NOTE_MOBILE', 'MOBILE_LONG_2', 'NOTE_MOBILE_2', 'MOBILE_LONG_3', 'NOTE_MOBILE_3', 'NOTE_FAXNR', 'FAXNR_LONG_2', 'NOTE_FAXNR_2', 'FAXNR_LONG_3', 'NOTE_FAXNR_3', 'NOTE_SMTP', 'SMTP_ADDR_2', 'NOTE_SMTP_2', 'SMTP_ADDR_3', 'NOTE_SMTP_3', 'URI_TYP', 'URI_ADDR', 'NOTE_URI', 'SPERR', 'SPERM']

#fields not to be considered in the related sheet for customer-extended existing record by new org. template (there is only a check about these fields are blank)
mf_customer_extend_company_data = ['ZWELS_08', 'ZWELS_09', 'ZWELS_10']
mf_customer_extend_sales_data = ['KVGR5']

#fields not to be considered in the related sheet for customer/vendor open items template (there is only a check about these fields are blank)
mf_bp_open_items = ['ZBD1T', 'ZBD1P', 'ZBD2T', 'ZBD2P', 'ZBD3T', 'SKFBT', 'ACSKT']

mf_gl_open_items = ['MWSKZ', 'TXJCD', 'NPLNR', 'COPA_KDGRP', 'COPA_BRSCH', 'COPA_KMLAND', 'COPA_ARTNR', 'COPA_KDPOS', 'COPA_KSTRG', 'COPA_PPRCTR', 'COPA_SERVICE_DOC_ID', 'COPA_SERVICE_DOC_ITEM_ID', 'COPA_SERVICE_DOC_TYPE']

#list of technical name fields, for which the maximum length is not well specified in the template, and so it needs a correction in the check
migration_file_1_max_digits = ['TAXKD', #CUSTOMER
                               'REPRF', 'WT_SUBJCT', 'WEBRE', #VENDOR
                               'UMSKZ', 'ZLSCH', 'ZLSPR'] #CUSTOMER OI

migration_file_2_max_digits = ['AFABE', #ASSET
                               'WITHT', 'WT_WITHCD', 'VTWEG', 'SPART', 'KDGRP', 'KONDA', 'KALKS', 'LPRIO', 'VSBED', 'KTGRD', 'PARVW', #CUSTOMER
                               'MWSKZ', 'WT_TYPE', 'WT_CODE'] #CUSTOMER OI

migration_file_3_max_digits = ['MEINS', 'ASSETTRTYP', #ASSET
                               'COUNTRY', 'REGION', 'INCO1', 'ALAND', 'BANKS', #CUSTOMER
                               'EKGRP', #VENDOR
                               'PROVZ'] #BANK

migration_file_4_max_digits = ['BUKRS', 'WERKS', 'GSBER', 'EVALGROUP1', 'EVALGROUP2', 'EVALGROUP3', 'EVALGROUP4', 'AFASL', #ASSET
                               'BU_GROUP', 'KTOKD', 'FRGRP', 'ZTERM', 'MAHNA', 'VKORG', 'VKBUR', 'TATYP', 'TAXTYPE', #CUSTOMER
                               'FRGRP', 'ZTERM1', 'EKORG', 'BSTAE', #VENDOR
                               'FKBER'] #CUSTOMER OI

migration_file_5_max_digits = ['CURRENCY', #ASSET
                               'HBKID', 'WAERS', #CUSTOMER
                               'HWAER', 'HWAE2', 'HWAE3'] #CUSTOMER OI

migration_file_6_max_digits = ['VBUND', 'BZIRK', #CUSTOMER
                               'RASSC'] #GL OI

migration_file_7_max_digits = ['BP_ROLE'] #CUSTOMER

migration_file_8_max_digits = ['ANLKL', 'EVALGROUP5', #ASSET
                               'PSPNR'] #CUSTOMER OI

migration_file_10_max_digits = ['KOSTL', 'VENDOR_NO', #ASSET
                                'KUNNR', 'AKONT', 'FDGRV', 'ZWELS_01', 'ZWELS_02', 'ZWELS_03', 'ZWELS_04', 'KUNN2', #CUSTOMER
                                'LIFNR', #VENDOR
                                'GKONT', 'PRCTR', #CUSTOMER OI
                                'COPA_PRCTR'] #GL OI

migration_file_28_max_digits = ['INCO2'] #CUSTOMER

mf_partner_function_customer = ['AG', 'RE', 'RG', 'WE', 'ZM']

mf_partner_function_supplier = ['LF', 'WL', 'BA', 'RS', 'ZM']

#only for this country the postal code additional check is set
mf_postal_code_country = ['AD', 'CA', 'CZ', 'DE', 'ES' 'FR', 'GB', 'GR', 'IT', 'MT', 'NG', 'NL', 'PL', 'PT', 'SE', 'SK', 'US']

#only for this country the vat additional check is set
mf_vat_country = ['IT', 'NG', 'AD']

#only for this country the bank data additional check is set
mf_bank_country = ['IT', 'ES', 'BE', 'FR', 'NL', 'FI', 'LU', 'CH', 'GB', 'DE', 'IE', 'NG']

class Root ():
    def __init__(
        self,
        root_title:str,
        tk_or_toplevel: str = 'TK',
        root_geometry:str = '1050x600'
    ):
        if tk_or_toplevel == 'TK':
            self.root = TkinterDnD.Tk()
        else:
            self.root = tkinter.Toplevel()
        self.root.title(root_title)
        self.root.configure(bg = '#F0F8FF')

        icon_path = 'communication_assistance_help_support_service_information_icon_230472.ico'
        self.root.iconbitmap(icon_path)

        self.root.state('zoomed')

        self.root.geometry(root_geometry)

class Frame ():
    def __init__(
        self,
        root: None,
        pack_or_grid: str = 'G',
        left_or_right: str = 'O',
        background: str = '#F0F8FF',
        column: int = 0,
        row: int = 0,
        sticky: str = '',
        col_span: int = 1,
        row_span: int = 1
    ):
        self.frame = tkinter.Frame(root, background = background)
        if pack_or_grid.upper() == 'P':
            if left_or_right.upper() == 'R':
                self.frame.pack(fill = tkinter.BOTH, expand = tkinter.TRUE, side = 'right')
            elif left_or_right.upper() == 'L':
                self.frame.pack(fill = tkinter.BOTH, expand = tkinter.TRUE, side = 'left')
            else:
                self.frame.pack(fill = tkinter.BOTH, expand = tkinter.TRUE)
        else:
            self.frame.grid(column = column, row = row, sticky = sticky, columnspan = col_span, rowspan = row_span)
            self.frame.columnconfigure(0, weight=1)  # Ensure column 0 expands

class Button ():
    def __init__ (
        self,
        frame: None,
        text: str = '',
        command: None = '',
        image: None = '',
        width: int = 0,
        height:int = 0,
        dimension: int = 12,
        x: int = 0,
        y: int = 0,
        anchor: str = 'nw',
        bordermode: str = 'inside'
    ):
        if image != '':
            img = tkinter.PhotoImage(file=image)
            self.button = tkinter.Button(frame, image = img, command = command)
        else:
            self.button = tkinter.Button(frame, text = text, command = command, width = width, height = height, background = '#D8E6EC', font = ('Calibri', dimension, 'bold'))
        self.button.place (x = x, y = y, anchor = anchor, bordermode = bordermode)

class Label ():
    def __init__ (
        self,
        frame: None,
        text: str,
        dimension: int = 13,
        weight = 'normal',
        justify = tkinter.LEFT,
        foreground: str = 'black',
        wraplength: int = 1000,
        x: int = 0,
        y: int = 0,
        anchor: str = 'nw',
        bordermode: str = 'inside'        
    ):
        self.label = tkinter.Label(frame, text = text, font = ('Calibri', dimension, weight), justify = justify, background = '#F0F8FF', foreground = foreground, wraplength = wraplength)
        self.label.place (x = x, y = y, anchor = anchor, bordermode = bordermode)

class Entry ():
    def __init__ (
        self,
        frame: None,
        width:int = 0,
        x: int = 0,
        y: int = 0,
    ):
        self.stringvar = tkinter.StringVar()
        self.entry = tkinter.Entry(frame, text = self.stringvar, width = width)
        self.entry.place (x = x, y = y)

class Combobox ():
    def __init__ (
        self,
        frame: None,
        command = '',
        values = '',
        width: int = 12,
        x: int = 0,
        y: int = 0,
        anchor: str = 'nw',
        bordermode: str = 'inside'
    ):
        self.text = tkinter.StringVar()
        self.combobox = ttk.Combobox(frame, textvariable = self.text, state = 'readonly', values = values, width = width)
        self.combobox.place (x = x, y = y, anchor = anchor, bordermode = bordermode)
        self.combobox.bind("<<ComboboxSelected>>", command)

class Checkbox ():
    def __init__(
        self,
        frame: None,
        text: str = '',
        command = '',
        x: int = 0,
        y: int = 0,
        anchor: str = 'nw',
        bordermode: str = 'inside'
    ):
        self.variable = tkinter.IntVar()
        self.checkbox = tkinter.Checkbutton(frame, text = text, variable = self.variable, command = command, background = '#F0F8FF')
        self.checkbox.place (x = x, y = y, anchor = anchor, bordermode = bordermode)

class TextEntry ():
    def __init__(
        self, 
        frame,
        text_height: int,
        text_width: int,
        height: int,
        width: int,
        entry_path: Entry,
        x: int = 0,
        y: int = 0
    ):
        self.entry_path = entry_path

        # Create a Text widget to display file content
        self.text_entry = tkinter.Text(frame, wrap="word", height=text_height, width=text_width)
        self.text_entry.place (x = x, y = y, height = height, width = width)

        # Add a label overlay on the Text widget
        self.drag_label = tkinter.Label(frame, text="Drag and drop a .txt file here", bg="white", fg="black")
        self.drag_label.place(in_=self.text_entry, relx=0.5, rely=0.5, anchor="center")

        # Bind the drop event to the Text widget
        self.text_entry.drop_target_register(DND_FILES)
        self.text_entry.dnd_bind("<<Drop>>", self.on_drop)

        self.y_scrollbar = ttk.Scrollbar(frame, orient='vertical', command=self.text_entry.yview)
        self.y_scrollbar.place(x = width + x, y = y, height = height)
        self.text_entry.configure(yscrollcommand=self.y_scrollbar.set)
    
    def on_drop(self, event):
        file_path = event.data.strip("{}")
        if file_path.endswith('.txt'):  # Only accept .txt files
            try:
                with open(file_path, 'r') as file:
                    content = file.read()
                    self.text_entry.delete(1.0, tkinter.END)  # Clear existing content
                    self.text_entry.insert(tkinter.END, content)  # Insert file content
                # Hide the drag-and-drop instruction label
                self.drag_label.place_forget()
            except Exception as e:
                self.text_entry.delete(1.0, tkinter.END)
                self.text_entry.insert(tkinter.END, f"Error reading file: {e}")
        else:
            self.text_entry.delete(1.0, tkinter.END)
            self.text_entry.insert(tkinter.END, "Please drop a valid .txt file.")
        self.entry_path.entry.config(state = 'normal')
        self.entry_path.entry.delete(0, tkinter.END)
        self.entry_path.entry.insert(0, file_path)
        self.entry_path.entry.config(state = 'disabled')
        
class MenuBar ():
    def __init__ (
        self,
        root: tkinter.Tk,
        first_label: str,
        second_label: str,
        third_label: str
    ): 
        self.menubar = tkinter.Menu (root)
        root.config (menu = self.menubar)
        self.main_menu_1 = tkinter.Menu (self.menubar, tearoff = 0)
        self.main_menu_2 = tkinter.Menu (self.menubar, tearoff = 0)
        self.main_menu_3 = tkinter.Menu (self.menubar, tearoff = 0)
        self.menubar.add_cascade (label = first_label, menu = self.main_menu_1)
        self.menubar.add_cascade (label = second_label, menu = self.main_menu_2)
        self.menubar.add_cascade (label = third_label, menu = self.main_menu_3)

class Treeview():
    def __init__(
        self,
        frame: tkinter.Frame,
        col_text: list,
        width_list: list,
        lst: list,
        dist: int = 0
    ):
        def select_all():
            # Get all item IDs in the Treeview
            item_ids = self.tree.get_children()

            # Select all items
            self.tree.selection_set(item_ids)
            on_copy(event=True)

        def on_copy(event):
            selected_items = self.tree.selection()
            if selected_items:
                copied_data = []
                for item_id in selected_items:
                    item_values = self.tree.item(item_id, 'values')
                    copied_data.append('\t'.join(str(value) for value in item_values))

                copied_text = '\n'.join(copied_data)
                pyperclip.copy(copied_text)

        self.frame = frame
        self.headers = col_text
        self.rows = lst


        # Configure Treeview style for bold, left-aligned headers
        style = ttk.Style(self.frame)
        style.configure(
            "Custom.Treeview.Heading",
            font=("Calibri", 12, "bold")
        )

        self.tree = ttk.Treeview(self.frame, columns=col_text, show='headings', height=len(lst), style="Custom.Treeview")
        for a in range(len(col_text)):
            self.tree.heading(col_text[a], text=col_text[a])
            self.tree.column(col_text[a], width=width_list[a])
        for b in lst:
            self.tree.insert('', tkinter.END, values=b)

        # Add a horizontal scrollbar
        self.x_scrollbar = ttk.Scrollbar(self.frame, orient='horizontal', command=self.tree.xview)
        self.x_scrollbar.place(x=10, y=710, width=1500)
        self.tree.configure(xscrollcommand=self.x_scrollbar.set)

        # Add a vertical scrollbar
        self.y_scrollbar = ttk.Scrollbar(self.frame, orient='vertical', command=self.tree.yview)
        self.y_scrollbar.place(x=1510, y=70 + dist, height=640 - dist)
        self.tree.configure(yscrollcommand=self.y_scrollbar.set)

        self.tree.place(x=10, y=70 + dist, width=1500, height=640 - dist)

        self.tree.bind("<Control-c>", on_copy)

        first_item = self.tree.get_children()[0]
        self.tree.selection_set(first_item)

        self.select_all_button = tkinter.Button(self.frame, text='Select All and Copy', command=select_all, background='#D8E6EC', font=('Calibri', 12, 'bold'))
        self.select_all_button.place(x=10, y=10, width=200, height=30)
    
    def export_to_excel(self, file_path: str, file_name: str, ebs: str = "NO"):
        # Extract data from Treeview
        data = [self.headers]

        for child in self.rows:
            data.append(child)

        # Create a Pandas DataFrame
        df = pd.DataFrame(data)

        current_datetime_str = datetime.now().strftime("%Y-%m-%d %H.%M.%S")

        father_path = file_path.rsplit("/", 1)[0]

        # Save the DataFrame to an Excel file
        file_path = father_path + "/" + file_name + f" - {current_datetime_str}.xlsx"
        df.to_excel(file_path, index=False, header=False)

        # Load the workbook to adjust column widths
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Apply bold font to the first and fourth rows
        bold_font = Font(bold=True)
        
        for row_num, row_cells in enumerate(sheet.iter_rows(), start=1):
            if ebs == "YES" and (row_num == 1 or row_num == 4):  # Make first and fourth rows bold
                for cell in row_cells:
                    cell.font = bold_font
            if row_num == 1: # Make first row bold
                for cell in row_cells:
                    cell.font = bold_font

        
        # Auto-adjust the column widths
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)  # Adjust for padding
            sheet.column_dimensions[column].width = adjusted_width

        # Save the workbook with adjusted column widths
        workbook.save(file_path)

        save_label = Label (
            frame = self.frame,
            text=f'The file was saved in {father_path}',
            foreground = '#006400',
            x = 10,
            y = 740
            )

class RadioButton_2 ():
    def __init__ (
        self,
        frame: tkinter.Frame,
        label_text: str = '',
        text_1: str = '',
        text_2: str = '',
        command = '',
        dimension: int = 8,
        x: int = 0,
        y: int = 0
    ):
        self.label_text = label_text
        self.text_1 = text_1
        self.text_2 = text_2
        self.label = Label (frame, text = label_text, x = x, y = y, dimension = dimension + 2)
        self.label.label.config(foreground='#191970')
        self.variable = tkinter.StringVar()
        self.radiobutton_1 = tkinter.Radiobutton (frame, text = text_1, variable = self.variable, value = text_1, command = command, font = ('Calibri', dimension), background = '#F0F8FF')
        self.radiobutton_1.place(x = x, y = y + 50)
        self.radiobutton_2 = tkinter.Radiobutton (frame, text = text_2, variable = self.variable, value = text_2, command = command, font = ('Calibri', dimension), background = '#F0F8FF')
        self.radiobutton_2.place(x = x, y = y + 100)

class RadioButton_3 (RadioButton_2):
    def __init__ (
        self,
        frame: tkinter.Frame,
        label_text: str = '',
        text_1: str = '',
        text_2: str = '',
        text_3: str = '',
        command = '',
        dimension: int = 9,
        x: int = 0,
        y: int = 0
    ):
        super ().__init__(
            frame,
            label_text,
            text_1,
            text_2,
            command,
            dimension,
            x,
            y
        )
        self.text_3 = text_3
        self.radiobutton_3 = tkinter.Radiobutton (frame, text = text_3, variable = self.variable, value = text_3, command = command, font = ('Calibri', dimension), background = '#F0F8FF')
        self.radiobutton_3.place(x = x, y = y + 150)
        self.text_input = tkinter.Text(frame, height = 37, width = 15)

class ScrollableFrame(tkinter.Frame):
    def __init__(self, master=None, **kwargs):
        tkinter.Frame.__init__(self, master, **kwargs)

        # Create a canvas and add it to the frame
        self.canvas = tkinter.Canvas(self, borderwidth=0, highlightthickness=0, background = '#F0F8FF')

        # Create a frame inside the canvas to hold the widgets
        self.inner_frame = tkinter.Frame(self.canvas, background = '#F0F8FF')

        # Add a horizontal scrollbar and link it to the canvas
        self.h_scrollbar = tkinter.Scrollbar(self, orient="horizontal", command=self.canvas.xview, background = '#F0F8FF')
        self.canvas.configure(xscrollcommand=self.h_scrollbar.set)

        # Pack the canvas and scrollbar into the frame
        self.canvas.pack(side="top", fill = tkinter.BOTH, expand = tkinter.TRUE)
        self.h_scrollbar.pack(side="bottom", fill="x")

        # Add the inner frame to the canvas
        self.canvas.create_window((0, 0), window=self.inner_frame, anchor="nw", height=680, width = 4000)

        # Configure the canvas to update the scroll region when the frame size changes
        self.inner_frame.bind("<Configure>", self.on_inner_frame_configure)

        # Bind the canvas to respond to mousewheel events for scrolling
        #self.canvas.bind_all("<MouseWheel>", self.on_mousewheel)

    def on_inner_frame_configure(self, event):
        # Update the scroll region to encompass the inner frame
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    #def on_mousewheel(self, event):
        # Handle mousewheel scrolling
    #    if event.delta:
    #        self.canvas.xview_scroll(int(-1 * (event.delta / 120)), "units")

class Sheet ():
    def __init__ (
        self,
        frame: tkinter.Frame,
        tab: ttk.Notebook
    ):
        self.sheet = Checkbox (frame)
        self.main_frame = Frame (tab, 'P')
        self.scrollable_frame = ScrollableFrame(self.main_frame.frame)
        self.scrollable_frame.pack(side = "top", fill = tkinter.BOTH, expand = tkinter.TRUE)
        self.frame_tab = self.scrollable_frame.inner_frame
        self.field_1 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_2 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_3 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_4 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required') 
        self.field_5 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_6 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_7 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_8 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_9 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_10 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_11 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_12 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_13 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_14 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_15 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_16 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_17 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_18 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_19 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_20 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_21 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_22 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_23 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_24 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_25 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_26 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_27 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_28 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_29 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required') 
        self.field_30 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_31 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_32 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_33 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_34 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_35 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_36 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_37 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_38 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_39 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_40 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_41 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_42 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_43 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_44 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_45 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_46 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_47 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_48 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_49 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')
        self.field_50 = RadioButton_3 (self.frame_tab, text_1 = 'Mandatory', text_2 = 'Optional', text_3 = 'Not Required')

        self.field_list = [self.field_1, self.field_2, self.field_3, self.field_4, self.field_5, self.field_6, self.field_7, self.field_8, self.field_9, self.field_10,
                           self.field_11, self.field_12, self.field_13, self.field_14, self.field_15, self.field_16, self.field_17, self.field_18, self.field_19, self.field_20,
                           self.field_21, self.field_22, self.field_23, self.field_24, self.field_25, self.field_26, self.field_27, self.field_28, self.field_29, self.field_30,
                           self.field_31, self.field_32, self.field_33, self.field_34, self.field_35, self.field_36, self.field_37, self.field_38, self.field_39, self.field_40,
                           self.field_41, self.field_42, self.field_43, self.field_44, self.field_45, self.field_46, self.field_47, self.field_48, self.field_49, self.field_50]




#A = letter; N = number; X = number/letter
def postal_code_check (postal_code: str, country: str):
    error = ''
    if country == 'AD' and not re.compile(r'^AD\d{3}$').match(postal_code):
        error = 'Should be in format "ANA NAN"'
    elif country == 'CA' and not re.compile(r'^[A-Z]{1}\d{1}[A-Z]{1} \d{1}[A-Z]{1}\d{1}$').match(postal_code):
        error = 'Should be in format "ANA NAN"'
    elif (country == 'CZ' or country == 'GR' or country == 'SE' or country == 'SK') and not re.compile(r'^\d{3} \d{2}$').match(postal_code):
        error = 'Should be in format "NNN NN"'
    elif (country == 'DE' or country == 'IT' or country == 'FR' or country == 'ES') and not re.compile(r'^\d{5}$').match(postal_code):
        error = 'Should be in format "NNNNN"'
    elif country == 'GB' and len(postal_code) > 9:
        error = 'Should not have a length greater than 9'
    elif country == 'MT' and not re.compile(r'^[A-Z]{3} \d{4}$').match(postal_code):
        error = 'Should be in format "AAA NNNN"'
    elif country == 'NL' and not re.compile(r'^\d{4} [A-Z]{2}$').match(postal_code):
        error = 'Should be in format "NNNN AA"'
    elif country == 'PL' and not re.compile(r'^\d{2}-\d{3}$').match(postal_code):
        error = 'Should be in format "NN-NNN"'
    elif country == 'US' and not re.compile(r'^\d{5}-\d{4}$').match(postal_code):
        error = 'Should be in format "NNNNN-NNNN"'
    

    return error

def vat_check (tax_type: str, tax_number:str, country: str):
    error = ''
    if country == 'IT':
        if tax_type == 'IT0' and not re.compile(r'^IT\d{11}$').match(tax_number):
            error = f'For country {country} and tax type {tax_type}, the tax number should be in format "ITNNNNNNNNNNN"'
        elif tax_type == 'IT1' and not re.compile(r'^[A-Z]{6}\d{2}[A-Z]{1}\d{2}[A-Z]{1}\d{3}[A-Z]{1}$').match(tax_number):
            error = f'For country {country} and tax type {tax_type}, the tax number should be in format "AAAAAANNANNANNNA"'
        elif tax_type == 'IT2' and not re.compile(r'^\d{11}$').match(tax_number):
            error = f'For country {country} and tax type {tax_type}, the tax number should be in format "NNNNNNNNNNN"'
        if tax_type != 'IT0' and tax_type != 'IT1' and tax_type != 'IT2':
            error = f'For country {country} the tax type values admitted are "IT0", "IT1" and "IT2"'
    if country == 'NG':
        if tax_type != 'NG1' and tax_type != 'NG3' and tax_type != 'NG4':
            error = f'For country {country} the tax type values admitted are "NG1", "NG3" and "NG4"'

    return error

def bank_check (sheet: str, bank_country: str, bank_key: str, bank_acc_number: str = '', bank_cont_key: str = '', iban: str = ''):
    error = ''
    if bank_country == 'IT' or bank_country == 'FR' :
        if not re.compile(r'^\d{10}$').match(bank_key):
            error = f'For country {bank_country} the bank key should be in format "NNNNNNNNNN". '

    elif bank_country == 'ES' or bank_country == 'DE':
        if not re.compile(r'^\d{8}$').match(bank_key):
            error = f'For country {bank_country} the bank key should be in format "NNNNNNNN". '

    elif bank_country == 'BE' or bank_country == 'LU' or bank_country == 'AE':
        if not re.compile(r'^\d{3}$').match(bank_key):
            error = f'For country {bank_country} the bank key should be in format "NNN". '

    elif bank_country == 'FI' or bank_country == 'GB' or bank_country == 'IE':
        if not re.compile(r'^\d{6}$').match(bank_key):
            error = f'For country {bank_country} the bank key should be in format "NNNNNN". '

    elif bank_country == 'CH':
        if not re.compile(r'^\d{5}$').match(bank_key):
            error = f'For country {bank_country} the bank key should be in format "NNNNN". '
    
    if sheet == 'Bank Details':
        if bank_country == 'IT':
            if not re.compile(r'^[A-Z0-9]{12}$').match(bank_acc_number):
                error += f'For country {bank_country} the bank account number should be in format "XXXXXXXXXXXX". '
            if not re.compile(r'^[A-Z]{1}$').match(bank_cont_key):
                error += f'For country {bank_country} the bank control key should be in format "A". '
            if not re.compile(fr'^IT\d{{2}}{re.escape(bank_cont_key)}{re.escape(bank_key)}{re.escape(bank_acc_number)}$').match(iban):
                error += 'Based on other data the IBAN is not correct; check it using IBAN transaction in this program'
        
        elif bank_country == 'ES':
            if not re.compile(r'^\d{10}$').match(bank_acc_number):
                error += f'For country {bank_country} the bank account number should be in format "NNNNNNNNNN". '
            if not re.compile(r'^\d{2}$').match(bank_cont_key):
                error += f'For country {bank_country} the bank control key should be in format "NN". '
            if not re.compile(fr'^ES\d{{2}}{re.escape(bank_key)}{re.escape(bank_cont_key)}{re.escape(bank_acc_number)}$').match(iban):
                error += 'Based on other data the IBAN is not correct; check it using IBAN transaction in this program'
            
        elif bank_country == 'BE':
            if not re.compile(fr'^{re.escape(bank_key)}-\d{{7}}-{re.escape(bank_cont_key)}$').match(bank_acc_number):
                error += f'For country {bank_country} the bank account number should be in format "NNN-NNNNNNN-NN". '
            if bank_cont_key != 'nan':
                error += f'For country {bank_country} the bank control key should be blank. '
            if not re.compile(fr'^BE\d{{2}}{re.escape(bank_key)}\d{{9}}$').match(iban):
                error += 'Based on other data the IBAN is not correct; check it using IBAN transaction in this program'

        elif bank_country == 'FR':
            if not re.compile(r'^[A-Z0-9]{11}$').match(bank_acc_number):
                error += f'For country {bank_country} the bank account number should be in format "XXXXXXXXXXX". '
            if not re.compile(r'^\d{2}$').match(bank_cont_key):
                error += f'For country {bank_country} the bank control key should be in format "NN". '
            if not re.compile(fr'^FR\d{{2}}{re.escape(bank_key)}{re.escape(bank_acc_number)}{re.escape(bank_cont_key)}$').match(iban):
                error += 'Based on other data the IBAN is not correct; check it using IBAN transaction in this program'

        elif bank_country == 'NL':
            if not re.compile(r'^\d{10}$').match(bank_acc_number):
                error += f'For country {bank_country} the bank account number should be in format "NNNNNNNNNN". '
            if bank_cont_key != 'nan':
                error += f'For country {bank_country} the bank control key should be blank. '
            if not re.compile(fr'^NL\d{{2}}[A-Z]{{4}}{re.escape(bank_acc_number)}$').match(iban):
                error += 'Based on other data the IBAN is not correct; check it using IBAN transaction in this program'

        elif bank_country == 'FI':
            if not re.compile(r'^\d{7}$').match(bank_acc_number):
                error += f'For country {bank_country} the bank account number should be in format "NNNNNNN". '
            if not re.compile(r'^\d{1}$').match(bank_cont_key):
                error += f'For country {bank_country} the bank control key should be in format "N". '
            if not re.compile(fr'^FI\d{{2}}{re.escape(bank_key)}{re.escape(bank_acc_number)}{re.escape(bank_cont_key)}$').match(iban):
                error += 'Based on other data the IBAN is not correct; check it using IBAN transaction in this program'

        elif bank_country == 'LU':
            if not re.compile(r'^\d{13}$').match(bank_acc_number):
                error += f'For country {bank_country} the bank account number should be in format "NNNNNNNNNNNNN". '
            if bank_cont_key != 'nan':
                error += f'For country {bank_country} the bank control key should be blank. '
            if not re.compile(fr'^LU\d{{2}}{re.escape(bank_key)}{re.escape(bank_acc_number)}$').match(iban):
                error += 'Based on other data the IBAN is not correct; check it using IBAN transaction in this program'

        elif bank_country == 'CH':
            if not re.compile(r'^\d{12}$').match(bank_acc_number):
                error += f'For country {bank_country} the bank account number should be in format "NNNNNNNNNNNN". '
            if bank_cont_key != 'nan':
                error += f'For country {bank_country} the bank control key should be blank. '
            if not re.compile(fr'^CH\d{{2}}{re.escape(bank_key)}{re.escape(bank_acc_number)}$').match(iban):
                error += 'Based on other data the IBAN is not correct; check it using IBAN transaction in this program'

        elif bank_country == 'GB' or bank_country == 'IE':
            if not re.compile(r'^\d{8}$').match(bank_acc_number):
                error += f'For country {bank_country} the bank account number should be in format "NNNNNNNN". '
            if bank_cont_key != 'nan':
                error += f'For country {bank_country} the bank control key should be blank. '
            if not re.compile(fr'^{re.escape(bank_country)}\d{{2}}[A-Z]{{4}}{re.escape(bank_key)}{re.escape(bank_acc_number)}$').match(iban):
                error += 'Based on other data the IBAN is not correct; check it using IBAN transaction in this program'

        elif bank_country == 'DE':
            if not re.compile(r'^\d{10}$').match(bank_acc_number):
                error += f'For country {bank_country} the bank account number should be in format "NNNNNNNNNN". '
            if bank_cont_key != 'nan':
                error += f'For country {bank_country} the bank control key should be blank. '
            if not re.compile(fr'^DE\d{{2}}{re.escape(bank_key)}{re.escape(bank_acc_number)}$').match(iban):
                error += 'Based on other data the IBAN is not correct; check it using IBAN transaction in this program'

        elif bank_country == 'NG':
            if bank_cont_key != 'nan':
                error = f'For country {bank_country} the bank control key should be blank. '
            if iban != 'nan':
                error += f'For country {bank_country} the IBAN should be blank'
        
        elif bank_country == 'AE':
            if not re.compile(r'^\d{16}$').match(bank_acc_number):
                error += f'For country {bank_country} the bank account number should be in format "NNNNNNNNNNNNNNNN". '
            if bank_cont_key != 'nan':
                error += f'For country {bank_country} the bank control key should be blank. '
            if not re.compile(fr'^AE\d{{2}}{re.escape(bank_key)}{re.escape(bank_acc_number)}$').match(iban):
                error += 'Based on other data the IBAN is not correct; check it using IBAN transaction in this program'

    return error

ebs_mt940_text = "It's possible to upload a .txt file or paste the content in the text space, in order to let the program reading the Electronic Bank Statement file. During analysis the program organizes the header data above and the line-item one at the bottom. It will be present a specific bank transaction type description, if it is known. There is a check regarding opening balance, closing balance and all the transactions: if there is a green tick, the check is ok, otherwise is not ok, and you can track the difference in the last row of the report."

iban_text = "It's possible to paste a sequence of IBANs in the text space. During analysis, for the country set in the program, the program splits the IBAN code into SAP fields."

mf_text = """
With the Migration File functionality, it's possible to analyze the .xlsx file related to SAP migration in S4/Hana.
The program allow to upload the file and recognizes all the related errors, based on input values decided by user.
In order to avoid program crash, the analysis will regard a maximum of 30 sheets and 50 fields per sheet.
After the file uploading there are some steps to be followed:
1.
The user should tick the checkboxes related to Excel sheets to be considered in the analysis.
In some cases the program recognizes a specific template and it will be possible to select a specific mode, in order to have specific fields to be considered and specific checks. In case of "Generic" mode the first 30 sheets and 50 fields per sheet will be considered.
2.
The user can choose the field status, field by field. It's possible to choose among: 1)mandatory; 2)optional; 3)not required.
It is possible to set a default for the field status. If the column header cell in Excel is green or is present the '*', so the field will be considered as mandatory in that sheet. If the column header cell in Excel is yellow or is present the '+', so the field will be considered as optional in that sheet.
3.
The user can choose the input values admitted for each field mandatory or optional. If the input field is left empty, so no input check will be made by the program.
It is possible to download a .xlsx template file. It is so possible to fill the Excel file and upload it in the program. In this way it's possible to fill all the input values only one time.
4.
The user can display all the errors recognized by the program. In the header it's possible to see the errors and warnings number.
Each error/warning is detailed with the sheet, the error/warning code, the row, the column name, the error/warning description.
It is possible to copy the errors/warnings table and paste it into a new Excel file, to manage the data.
It is also possible to change the file under analysis and do again the analysis with the same parameters, without repeating all the previous steps.
"""